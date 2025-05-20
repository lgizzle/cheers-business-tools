import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import json
from datetime import datetime

class SingleDealCalculator:
    """
    Single Product Deal Calculator to help determine whether to purchase
    smaller quantities or larger quantities (bulk) based on different price points
    and other business considerations.
    """

    def __init__(self):
        """Initialize the calculator"""
        pass

    def calculate_deal(self, params):
        """
        Calculate the deal metrics and provide recommendations

        Args:
            params: Dictionary containing input parameters
                - smaller_deal_qty: Number of cases in smaller order quantity
                - bulk_deal_qty: Number of cases in bulk order quantity
                - price_per_bottle_smaller: Price per bottle when ordering smaller quantity
                - price_per_bottle_bulk: Price per bottle when ordering bulk quantity
                - annual_sales_volume: Expected annual sales in cases
                - vendor_terms: Days before payment is due
                - bottles_per_case: Number of bottles in each case

        Returns:
            Dictionary with calculation results and recommendations
        """
        # Extract parameters
        smaller_deal_qty = float(params.get('smaller_deal_qty', 0))
        bulk_deal_qty = float(params.get('bulk_deal_qty', 0))
        price_per_bottle_smaller = float(params.get('price_per_bottle_smaller', 0))
        price_per_bottle_bulk = float(params.get('price_per_bottle_bulk', 0))
        annual_sales_volume = float(params.get('annual_sales_volume', 0))
        vendor_terms = float(params.get('vendor_terms', 0))
        bottles_per_case = float(params.get('bottles_per_case', 0))

        # Calculate daily sales rate
        daily_sales_rate = annual_sales_volume / 365

        # Calculate leftover inventory after terms
        leftover_smaller = max(smaller_deal_qty - (vendor_terms * daily_sales_rate), 0)
        leftover_bulk = max(bulk_deal_qty - (vendor_terms * daily_sales_rate), 0)

        # Calculate average cash tied up (assuming linear sales rate)
        avg_cash_tied_smaller = 0.5 * leftover_smaller * price_per_bottle_smaller * bottles_per_case
        avg_cash_tied_bulk = 0.5 * leftover_bulk * price_per_bottle_bulk * bottles_per_case
        extra_cash_tied_up = avg_cash_tied_bulk - avg_cash_tied_smaller

        # Calculate savings from bulk discount
        total_savings = bulk_deal_qty * bottles_per_case * (price_per_bottle_smaller - price_per_bottle_bulk)

        # Calculate how long cash is tied up
        days_tied_up = 0 if daily_sales_rate == 0 else leftover_bulk / daily_sales_rate

        # Calculate return on investment
        if extra_cash_tied_up == 0:
            roi = float('inf')
            roi_text = "Infinite"
            annualized_roi = float('inf')
            annualized_roi_text = "Infinite"
        else:
            roi = total_savings / extra_cash_tied_up
            roi_text = f"{roi:.2%}"
            annualized_roi = 0 if days_tied_up == 0 else roi * (365 / days_tied_up)
            annualized_roi_text = f"{annualized_roi:.2%}"

        # Generate recommendation
        if isinstance(annualized_roi, float) and annualized_roi > 0.6:
            recommendation = "Take the bulk deal - excellent return on investment!"
        elif isinstance(annualized_roi, float) and annualized_roi > 0.05:
            recommendation = "Consider the bulk deal - decent return."
        else:
            recommendation = "Stick with smaller deal - return is not worth the extra cash tied up."

        # Context for the recommendation
        roi_context = f"Your annual return on the extra investment is {annualized_roi_text}."
        cash_flow_impact = f"You will need to invest an extra ${extra_cash_tied_up:,.2f} in inventory but will save ${total_savings:,.2f} per order."

        # Build result dictionary
        results = {
            'daily_sales_rate': daily_sales_rate,
            'leftover_smaller': leftover_smaller,
            'leftover_bulk': leftover_bulk,
            'avg_cash_tied_smaller': avg_cash_tied_smaller,
            'avg_cash_tied_bulk': avg_cash_tied_bulk,
            'extra_cash_tied_up': extra_cash_tied_up,
            'total_savings': total_savings,
            'days_tied_up': days_tied_up,
            'roi': roi_text,
            'annualized_roi': annualized_roi_text,
            'recommendation': recommendation,
            'roi_context': roi_context,
            'cash_flow_impact': cash_flow_impact
        }

        return results

    def generate_report(self, params, file_path):
        """
        Generate an Excel report with the deal calculation results

        Args:
            params: Dictionary containing input parameters
            file_path: Path to save the Excel file

        Returns:
            Path to the generated Excel file
        """
        # Calculate results
        results = self.calculate_deal(params)

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Single-Product Deal Calculator"

        # Add title
        ws['A1'] = "CHEERS LIQUOR MART"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = "SINGLE PRODUCT DEAL BUYING CALCULATOR"
        ws['A2'].font = Font(size=12, bold=True)
        ws.merge_cells('A2:C2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Add header for input section
        ws['A4'] = "Input Variables"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = "Amount"
        ws['B4'].font = Font(bold=True)
        ws['C4'] = "Input Explanations"
        ws['C4'].font = Font(bold=True)

        # Fill input section
        input_rows = [
            ("Smaller Deal (cases ordered)", params['smaller_deal_qty'], "Number of cases in your smaller order quantity"),
            ("Bulk Deal (cases ordered)", params['bulk_deal_qty'], "Number of cases in your bulk order quantity"),
            ("Price per Bottle (Smaller Deal)", params['price_per_bottle_smaller'], "Price per bottle when ordering the smaller quantity"),
            ("Price per Bottle (Bulk Deal)", params['price_per_bottle_bulk'], "Price per bottle when ordering the bulk quantity (should be lower)"),
            ("Estimated Sales Volume (cases/year)", params['annual_sales_volume'], "How many cases you expect to sell in a year"),
            ("Vendor Terms (days before payment is due)", params['vendor_terms'], "How many days before you need to pay the vendor"),
            ("Bottles in Each Case", params['bottles_per_case'], "Number of bottles contained in each case")
        ]

        row = 5
        for label, value, explanation in input_rows:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=3).value = explanation
            row += 1

        # Add header for results section
        ws['A13'] = "Calculation Results"
        ws['A13'].font = Font(bold=True)
        ws.merge_cells('A13:C13')

        # Fill results section
        result_rows = [
            ("How Fast It Sells (cases per day)", results['daily_sales_rate'], "Annual sales volume divided by 365 days"),
            ("Leftover Inventory After Terms (Smaller Deal)", results['leftover_smaller'], "Cases remaining after expected sales during payment term"),
            ("Leftover Inventory After Terms (Bulk Deal)", results['leftover_bulk'], "Cases remaining after expected sales during payment term"),
            ("Avg $ Tied Up After Terms (Smaller Deal)", results['avg_cash_tied_smaller'], "Average cash value of remaining inventory (assuming linear sales)"),
            ("Avg $ Tied Up After Terms (Bulk Deal)", results['avg_cash_tied_bulk'], "Average cash value of remaining inventory (assuming linear sales)"),
            ("Extra Cash You're Tying Up with Bulk", results['extra_cash_tied_up'], "Extra cash tied up with bulk deal vs smaller deal"),
            ("Total Savings from Bulk Price", results['total_savings'], "Total savings from bulk discount"),
            ("How Long Your Cash is Tied Up (days)", results['days_tied_up'], "How many days it takes to sell through remaining inventory"),
            ("Return on Extra Investment", results['roi'], "Return on extra investment (savings divided by extra cash tied up)"),
            ("Annualized Return", results['annualized_roi'], "Annualized return (if you could repeat this cycle all year)")
        ]

        row = 14
        for label, value, explanation in result_rows:
            ws.cell(row=row, column=1).value = label

            # Format numeric values
            if isinstance(value, float):
                if 'Return' in label:
                    ws.cell(row=row, column=2).value = value
                else:
                    ws.cell(row=row, column=2).value = value
                    ws.cell(row=row, column=2).number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=2).value = value

            ws.cell(row=row, column=3).value = explanation
            row += 1

        # Add recommendation section
        ws['A25'] = "Decision Recommendation"
        ws['A25'].font = Font(bold=True)
        ws.merge_cells('A25:C25')

        ws['A26'] = "Recommendation:"
        ws['B26'] = results['recommendation']
        ws['B26'].font = Font(bold=True)
        if "excellent" in results['recommendation']:
            ws['B26'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif "decent" in results['recommendation']:
            ws['B26'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            ws['B26'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws['A27'] = "Context:"
        ws['B27'] = results['roi_context']

        ws['A28'] = "Cash Flow Impact:"
        ws['B28'] = results['cash_flow_impact']

        # Set column widths
        column_widths = {'A': 40, 'B': 25, 'C': 50}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Format currency cells
        currency_rows = [7, 8, 17, 18, 19, 20]
        for row in currency_rows:
            ws.cell(row=row, column=2).number_format = '$#,##0.00'

        # Save the workbook
        wb.save(file_path)

        return file_path
