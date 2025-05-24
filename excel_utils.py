from openpyxl.styles import Font, PatternFill, Alignment

def get_title_font():
    return Font(name='Calibri', size=16, bold=True)

def get_header_font():
    return Font(name='Calibri', size=12, bold=True, color='FFFFFF')

def get_header_fill():
    return PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

def get_money_format():
    return '#,##0.00'

def get_percent_format():
    return '0.00%'

def apply_header_styles(ws, row=1, start_col=1, end_col=None):
    """Apply header font, fill, and alignment to a row."""
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = Alignment(horizontal='center')
