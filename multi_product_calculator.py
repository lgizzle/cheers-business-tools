"""
Multi-Product Buying Calculator.
Analyzes the ROI of purchasing multiple products at bulk discount pricing.
Core backend implementation for the Python Flask app.
"""

import os
import json
import logging
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from scenario_utils import load_scenario_file, save_scenario_file, delete_scenario_file, list_scenario_files
from logging_utils import setup_logging
from excel_utils import get_title_font, get_header_font, get_header_fill, get_money_format, get_percent_format, apply_header_styles
from api_utils import validate_numeric
import math

# Set up logging
setup_logging('logs/calculator.log')

class MultiProductBuyingCalculator:
    """
    Multi-Product Buying Calculator.
    Analyzes the ROI of purchasing multiple related products at bulk discount pricing.
    """

    def __init__(self):
        """Initialize the calculator with default parameters."""
        self.scenarios_dir = "scenarios/multi_product"
        os.makedirs(self.scenarios_dir, exist_ok=True)
        self.logger = logging.getLogger(__name__)

    def compute_line_item_roi(self, product, params):
        """
        Compute ROI metrics for a single product line item using linear depletion model.

        ROI = Total Savings ÷ ΔInvestment, where:
        - Total Savings = Q₂ × B × (P₁-P₂)
        - Cases sold during payment terms
        - Cases left after payment terms
        - Average dollar value of cases left (assuming linear depletion)
        - ΔInvestment = avg_dollar_bulk - avg_dollar_small

        Args:
            product (dict): Product data
            params (dict): Calculation parameters

        Returns:
            dict: ROI metrics
        """
        try:
            # Extract product data
            annual_cases = product.get('annual_cases', 0)           # Annual velocity in cases
            bottles_per_case = product.get('bottles_per_case', 0)    # B: Bottles per case
            price_small = product.get('current_price', 0)             # P₁: Price per bottle in small deal
            price_bulk = product.get('bulk_price', 0)               # P₂: Price per bottle in bulk deal
            bulk_cases = product.get('bulk_quantity', 0)               # Q₂: Cases purchased in bulk
            on_hand_cases = product.get('on_hand', 0)          # Current inventory in cases

            # Extract parameters
            payment_terms_days = params.get('paymentTermsDays', 30)
            min_days_stock = params.get('minDaysStock', 30)
            deal_size_cases = params.get('dealSizeCases', 60)      # Bulk deal minimum in cases
            # Handle both old and new parameter names for backward compatibility
            small_deal_minimum = params.get('smallDealCases', params.get('smallDealMinimum', 30)) # Small deal minimum in cases

            # Edge case: No velocity
            if annual_cases <= 0:
                return {
                    "avgInvSmall": 0,
                    "avgInvBulk": 0,
                    "deltaInvestment": 0,
                    "savings": 0,
                    "roi": 0,
                    "annualizedRoi": 0,
                    "dealCyclesPerYear": 0,
                    "annualROIMultiplier": 0,
                    "error": "Zero velocity"
                }

            # Edge case: No bulk cases
            if bulk_cases <= 0:
                return {
                    "avgInvSmall": 0,
                    "avgInvBulk": 0,
                    "deltaInvestment": 0,
                    "savings": 0,
                    "roi": 0,
                    "annualizedRoi": 0,
                    "smallDealCases": 0,
                    "dealCyclesPerYear": 0,
                    "annualROIMultiplier": 0
                }

            # Calculate daily velocity
            daily_cases = annual_cases / 365  # V: Daily velocity in cases

            # Calculate Small Deal quantity - purely proportional, not influenced by days stock
            # Q₁ = ceiling(Q₂ × SmallDealMin/BulkDealMin)
            small_deal_cases = math.ceil(bulk_cases * (small_deal_minimum / deal_size_cases))

            # Days-of-stock baseline is only used for validation/warnings, not for Small Deal calculation
            days_qty = math.ceil(daily_cases * min_days_stock)

            # 1. Calculate Total Savings: Q₂ × B × (P₁-P₂)
            savings_per_bottle = price_small - price_bulk
            total_savings = bulk_cases * bottles_per_case * savings_per_bottle

            # 2. Calculate cases sold during payment terms period
            cases_sold_during_terms = daily_cases * payment_terms_days

            # 3. Calculate cases left over after payment terms for both deals
            small_cases_left = max(0, small_deal_cases - cases_sold_during_terms)
            bulk_cases_left = max(0, bulk_cases - cases_sold_during_terms)

            # 4. Calculate dollar value of cases left over
            small_dollar_value = small_cases_left * price_small * bottles_per_case
            bulk_dollar_value = bulk_cases_left * price_bulk * bottles_per_case

            # 5. Calculate average dollar value of cases left over (assuming linear depletion)
            avg_dollar_small = small_dollar_value / 2
            avg_dollar_bulk = bulk_dollar_value / 2

            # 6. Calculate delta investment
            delta_investment = avg_dollar_bulk - avg_dollar_small

            # 7. Calculate ROI
            roi = total_savings / delta_investment if delta_investment > 0 else 0

            # Calculate depletion periods for reference (total time to deplete bulk inventory)
            days_to_deplete_bulk = bulk_cases / daily_cases if daily_cases > 0 else float('inf')

            # Calculate days at risk (after payment terms expire) for annualized ROI
            # This is the period where capital is actually tied up and at risk
            days_at_risk = bulk_cases_left / daily_cases if daily_cases > 0 and bulk_cases_left > 0 else 0

            # Calculate annualized ROI based only on after-terms exposure period
            # Formula: ROI × (365 ÷ DaysAtRisk)
            if days_at_risk > 0:
                annualized_roi = roi * (365 / days_at_risk)
            else:
                # If no capital at risk (all inventory sold during payment terms), ROI is already annualized
                annualized_roi = roi

            # Calculate stock days for warning check
            total_stock_days = (on_hand_cases + bulk_cases) / daily_cases if daily_cases > 0 else float('inf')

            # Calculate deal cycles per year (raw inventory turnover, unadjusted)
            # Formula: Annual Cases / Average Inventory
            avg_inventory_cases = (on_hand_cases + bulk_cases) / 2
            deal_cycles_per_year = annual_cases / avg_inventory_cases if avg_inventory_cases > 0 else 0

            # Calculate annual ROI multiplier (after-terms) - this is what's used for portfolio calculations
            # This reflects the true annual cash-on-cash return pace after payment terms expire
            if days_at_risk > 0:
                annual_roi_multiplier = 365 / days_at_risk
            else:
                annual_roi_multiplier = 0

            # Set warning flag if minimum days stock not met, but don't zero out ROI
            warning_message = None
            if total_stock_days < min_days_stock:
                warning_message = f"Insufficient stock: {total_stock_days:.1f} days vs {min_days_stock} days required"

            # For debugging and verification purposes
            debug_info = {
                "dailyVelocity": daily_cases,
                "Q2": bulk_cases,
                "Q1": small_deal_cases,
                "daysQty": days_qty,
                "minDaysStock": min_days_stock,
                "smallDealMin": small_deal_minimum,
                "bulkDealMin": deal_size_cases,
                "casesSoldDuringTerms": cases_sold_during_terms,
                "smallCasesLeft": small_cases_left,
                "bulkCasesLeft": bulk_cases_left,
                "smallDollarValue": small_dollar_value,
                "bulkDollarValue": bulk_dollar_value,
                "avgDollarSmall": avg_dollar_small,
                "avgDollarBulk": avg_dollar_bulk,
                "savings": total_savings,
                "deltaInv": delta_investment,
                "avgInventoryCases": avg_inventory_cases,
                "dealCyclesPerYear": deal_cycles_per_year,
                "annualROIMultiplier": annual_roi_multiplier,
                "daysToDepleteBulk": days_to_deplete_bulk,
                "daysAtRisk": days_at_risk
            }

            result = {
                "smallDealCases": float(small_deal_cases),
                "avgInvSmall": float(avg_dollar_small),
                "avgInvBulk": float(avg_dollar_bulk),
                "deltaInvestment": float(delta_investment),
                "savings": float(total_savings),
                "roi": float(roi),
                "annualizedRoi": float(annualized_roi),
                "daysToDepleteBulk": float(days_to_deplete_bulk),
                "totalStockDays": float(total_stock_days),
                "dealCyclesPerYear": float(deal_cycles_per_year),
                "annualROIMultiplier": float(annual_roi_multiplier),
                "debug": debug_info
            }

            if warning_message:
                result["warning"] = warning_message

            return result

        except Exception as e:
            self.logger.error(f"Error computing ROI: {str(e)}")
            return {
                "deltaInvestment": 0,
                "savings": 0,
                "roi": 0,
                "annualizedRoi": 0,
                "dealCyclesPerYear": 0,
                "annualROIMultiplier": 0,
                "error": f"Calculation error: {str(e)}"
            }

    def allocate_based_on_need(self, products, deal_size_cases, min_days_stock=30):
        """
        Allocate deal cases based on product needs, taking into account on-hand inventory.

        This function allocates cases based on how many additional cases each product
        needs to reach the minimum days of stock requirement, rather than just annual sales.
        If the total need is less than the deal size, the remaining cases are allocated
        proportionally based on annual sales.

        Args:
            products (list): List of product dictionaries
            deal_size_cases (int): Total number of cases in the deal
            min_days_stock (int): Minimum days of stock required

        Returns:
            list: Updated products with bulkCases assigned
        """
        try:
            # Ensure deal_size_cases is an integer
            deal_size_cases = int(deal_size_cases)

            # Make a copy of products to avoid modifying the original
            allocated_products = []

            # Calculate daily depletion rates and current days of stock
            products_with_need = []
            total_annual_cases = 0

            for product in products:
                product_copy = product.copy()

                # Calculate daily depletion rate
                annual_cases = product.get('annual_cases', 0)
                total_annual_cases += annual_cases

                if annual_cases <= 0:
                    # Skip products with no sales velocity
                    product_copy['bulk_quantity'] = 0
                    allocated_products.append(product_copy)
                    continue

                daily_cases = annual_cases / 365

                # Calculate current days of stock
                on_hand_cases = product.get('on_hand', 0)
                current_days_stock = on_hand_cases / daily_cases if daily_cases > 0 else float('inf')

                # Calculate additional cases needed to reach minimum days of stock
                additional_cases_needed = max(0, (min_days_stock - current_days_stock) * daily_cases)

                products_with_need.append({
                    'product': product_copy,
                    'daily_cases': daily_cases,
                    'current_days_stock': current_days_stock,
                    'additional_cases_needed': additional_cases_needed,
                    'annual_cases': annual_cases
                })

            # Calculate total need across all products
            total_need = sum(p['additional_cases_needed'] for p in products_with_need)

            # If total need is zero, fall back to proportional allocation by annual sales
            if total_need <= 0:
                self.logger.info("No additional cases needed based on minimum days stock - falling back to proportional allocation")
                return self.allocate_proportional(products, deal_size_cases)

            # STEP 1: First, allocate based on need proportion
            remaining_cases = deal_size_cases

            for product_info in products_with_need:
                product = product_info['product']
                need = product_info['additional_cases_needed']

                # Calculate raw case allocation based on need proportion
                if total_need > 0:
                    need_proportion = need / total_need
                    allocated_cases = min(need, int(deal_size_cases * need_proportion))
                else:
                    allocated_cases = 0

                # Ensure allocated_cases is an integer
                product['bulk_quantity'] = int(allocated_cases)
                remaining_cases -= int(allocated_cases)

            # STEP 2: If we've allocated all cases based on need and have leftover,
            # distribute the remaining cases proportionally based on annual sales
            if remaining_cases > 0:
                self.logger.info(f"Need-based allocation used {deal_size_cases - remaining_cases} cases. Allocating remaining {remaining_cases} cases based on annual sales.")

                # Calculate proportional distribution for remaining cases
                if total_annual_cases > 0:
                    products_by_annual = sorted(products_with_need,
                                               key=lambda p: p['annual_cases'],
                                               reverse=True)

                    # Allocate remaining cases proportionally by annual sales
                    # Use fractional allocation to maximize fairness
                    annual_fractions = []
                    for product_info in products_by_annual:
                        annual_proportion = product_info['annual_cases'] / total_annual_cases
                        fractional_cases = remaining_cases * annual_proportion
                        int_cases = int(fractional_cases)
                        product_info['product']['bulk_quantity'] += int_cases
                        remaining_cases -= int_cases

                        annual_fractions.append({
                            'product_info': product_info,
                            'fraction': fractional_cases - int_cases
                        })

                    # Distribute any remaining individual cases by highest fraction
                    annual_fractions.sort(key=lambda x: x['fraction'], reverse=True)
                    for i in range(min(int(remaining_cases), len(annual_fractions))):
                        annual_fractions[i]['product_info']['product']['bulk_quantity'] += 1

            # Final sanity check - ensure we've allocated exactly the requested number of cases
            total_allocated = sum(p['product']['bulk_quantity'] for p in products_with_need)
            if total_allocated != deal_size_cases:
                self.logger.warning(f"Allocation discrepancy: Target={deal_size_cases}, Actual={total_allocated}. Adjusting.")

                # If we're off, adjust the allocation of the product with the highest annual cases
                adjustment = deal_size_cases - total_allocated
                products_by_annual = sorted(products_with_need,
                                          key=lambda p: p['annual_cases'],
                                          reverse=True)

                for product_info in products_by_annual:
                    # Only increase allocation (if adjustment > 0) or decrease if product has enough cases
                    if adjustment > 0 or product_info['product']['bulk_quantity'] >= abs(adjustment):
                        product_info['product']['bulk_quantity'] += adjustment
                        break

            # Collect all products into the final result
            for product_info in products_with_need:
                # Final check to ensure all bulk_quantity are integers
                product_info['product']['bulk_quantity'] = int(product_info['product']['bulk_quantity'])
                allocated_products.append(product_info['product'])

            self.logger.info(f"Allocated {sum(p.get('bulk_quantity', 0) for p in allocated_products)} cases based on inventory needs and annual sales")
            return allocated_products

        except Exception as e:
            self.logger.error(f"Error in need-based allocation: {str(e)}")
            # Fall back to proportional allocation if there's an error
            self.logger.info("Falling back to proportional allocation due to error")
            return self.allocate_proportional(products, deal_size_cases)

    def allocate_proportional(self, products, deal_size_cases):
        """
        Allocate deal cases proportionally based on annual sales.

        Args:
            products (list): List of product dictionaries
            deal_size_cases (int): Total number of cases in the deal

        Returns:
            list: Updated products with bulkCases assigned
        """
        try:
            # Calculate total annual cases
            total_annual = sum(p.get('annual_cases', 0) for p in products)

            if total_annual <= 0:
                raise ValueError("Total annual cases must be greater than zero")

            # Make a copy of products to avoid modifying the original
            allocated_products = []
            remainder_info = []

            # Calculate raw case allocation and initial floor allocation
            for product in products:
                annual_cases = product.get('annual_cases', 0)
                raw_cases = deal_size_cases * (annual_cases / total_annual)
                bulk_cases = int(raw_cases)  # Floor
                remainder = raw_cases - bulk_cases  # Fractional part

                product_copy = product.copy()
                product_copy['bulk_quantity'] = bulk_cases
                allocated_products.append(product_copy)
                remainder_info.append({
                    'index': len(allocated_products) - 1,
                    'remainder': remainder
                })

            # Calculate leftover cases
            allocated_cases = sum(p.get('bulk_quantity', 0) for p in allocated_products)
            leftover = deal_size_cases - allocated_cases

            # Sort by remainder in descending order for leftover distribution
            remainder_info.sort(key=lambda x: x['remainder'], reverse=True)

            # Distribute leftover cases
            for i in range(int(min(leftover, len(remainder_info)))):
                idx = remainder_info[i]['index']
                allocated_products[idx]['bulk_quantity'] += 1

            return allocated_products

        except Exception as e:
            self.logger.error(f"Error in proportional allocation: {str(e)}")
            raise ValueError(f"Allocation error: {str(e)}")

    def run_iterations(self, products, params):
        """
        Run optimization iterations to improve ROI through case swaps.

        Args:
            products (list): List of product dictionaries with initial bulkCases
            params (dict): Calculation parameters

        Returns:
            dict: Updated products and iteration history
        """
        try:
            iterations = params.get('iterations', 'auto')
            min_days_stock = params.get('minDaysStock', 30)
            payment_terms_days = params.get('paymentTermsDays', 30)

            history = []
            iteration_count = 0
            max_iterations = 100 if iterations == 'auto' else int(iterations)
            improved = True

            # Deep copy products to avoid modifying the original
            current_products = [p.copy() for p in products]

            # Initial portfolio ROI calculation
            portfolio_metrics = self.calculate_portfolio_roi(current_products, params)
            portfolio_annualized_roi = portfolio_metrics['roi'] * portfolio_metrics['annualROIMultiplier']
            self.logger.info(f"Initial portfolio ROI: {portfolio_metrics['roi']:.4f}, Annualized: {portfolio_annualized_roi:.4f}")

            history.append({
                'iteration': 0,
                'totalROI': float(portfolio_metrics['roi']),
                'totalAnnualizedROI': float(portfolio_annualized_roi),
                'products': [p.copy() for p in current_products]
            })

            # Run iterations
            while improved and iteration_count < max_iterations:
                improved = False
                iteration_count += 1
                self.logger.info(f"Starting iteration {iteration_count}")

                # Find lowest and highest ROI products
                lowest_roi_product = None
                highest_roi_product = None
                lowest_roi = float('inf')
                highest_roi = float('-inf')

                # Calculate ROI for each product
                products_with_roi = []
                for product in current_products:
                    roi_metrics = self.compute_line_item_roi(product, params)
                    annualized_roi = roi_metrics.get('annualizedRoi', 0)  # Use annualized ROI for optimization

                    product_with_roi = product.copy()
                    product_with_roi['annualizedRoi'] = annualized_roi
                    products_with_roi.append(product_with_roi)

                    # Track lowest annualized ROI (only for products with bulk cases)
                    if annualized_roi < lowest_roi and product.get('bulk_quantity', 0) > 0:
                        lowest_roi = annualized_roi
                        lowest_roi_product = product

                    # Track highest annualized ROI (only for products with room for more)
                    annual_cases = product.get('annual_cases', 0)
                    if annualized_roi > highest_roi and product.get('bulk_quantity', 0) < annual_cases:
                        highest_roi = annualized_roi
                        highest_roi_product = product

                # Try a swap if we found candidates
                if lowest_roi_product and highest_roi_product:
                    self.logger.info(f"Attempting swap from {lowest_roi_product.get('product_name')} (Annualized ROI: {lowest_roi:.4f}) to {highest_roi_product.get('product_name')} (Annualized ROI: {highest_roi:.4f})")

                    # Create test copies
                    test_products = [p.copy() for p in current_products]

                    # Find indexes of the products
                    low_index = next(i for i, p in enumerate(test_products)
                                  if p.get('id') == lowest_roi_product.get('id'))
                    high_index = next(i for i, p in enumerate(test_products)
                                   if p.get('id') == highest_roi_product.get('id'))

                    # Perform the swap
                    test_products[low_index]['bulk_quantity'] -= 1
                    test_products[high_index]['bulk_quantity'] += 1

                    # Calculate new portfolio metrics (we no longer check minimum days stock here)
                    new_portfolio_metrics = self.calculate_portfolio_roi(test_products, params)
                    new_portfolio_annualized_roi = new_portfolio_metrics['roi'] * new_portfolio_metrics['annualROIMultiplier']
                    self.logger.info(f"New portfolio ROI after swap: {new_portfolio_metrics['roi']:.4f}, Annualized: {new_portfolio_annualized_roi:.4f}")

                    # Accept the swap if it improves the ANNUALIZED portfolio ROI
                    if new_portfolio_annualized_roi > portfolio_annualized_roi:
                        current_products = test_products
                        portfolio_metrics = new_portfolio_metrics
                        portfolio_annualized_roi = new_portfolio_annualized_roi
                        improved = True
                        self.logger.info(f"Swap accepted - portfolio annualized ROI improved to {portfolio_annualized_roi:.4f}")

                        # Record this iteration
                        history.append({
                            'iteration': iteration_count,
                            'totalROI': float(portfolio_metrics['roi']),
                            'totalAnnualizedROI': float(portfolio_annualized_roi),
                            'swapped': {
                                'from': lowest_roi_product.get('product_name', ''),
                                'to': highest_roi_product.get('product_name', '')
                            },
                            'products': [p.copy() for p in current_products]
                        })
                    else:
                        self.logger.info(f"Swap rejected - portfolio annualized ROI would decrease to {new_portfolio_annualized_roi:.4f}")
                else:
                    self.logger.info("No suitable candidates found for swapping")

                # Break if no improvement and we're in auto mode
                if not improved and iterations == 'auto':
                    self.logger.info(f"No improvement in iteration {iteration_count}, stopping optimization")
                    break

            self.logger.info(f"Optimization completed after {iteration_count} iterations. Final ROI: {portfolio_metrics['roi']:.4f}, Annualized: {portfolio_annualized_roi:.4f}")
            return {
                'products': current_products,
                'history': history,
                'totalIterations': iteration_count,
                'finalROI': float(portfolio_metrics['roi']),
                'finalAnnualizedROI': float(portfolio_annualized_roi)
            }

        except Exception as e:
            self.logger.error(f"Error in optimization: {str(e)}")
            raise ValueError(f"Optimization error: {str(e)}")

    def check_min_days_stock(self, product, min_days_stock):
        """
        Check if a product meets minimum days stock requirement.

        This function now returns True even if minimum days stock is not met,
        as we want to show warnings rather than blocking optimization.

        Args:
            product (dict): Product dictionary
            min_days_stock (int): Minimum days of stock required

        Returns:
            bool: True if minimum days stock requirement is met or if we're allowing
                 products below minimum stock (with warnings)
        """
        # Always return True - we now use warnings instead of hard requirements
        # This allows optimization to proceed even with products below minimum stock
        return True

    def calculate_portfolio_roi(self, products, params):
        """
        Calculate overall portfolio ROI using after-terms exposure methodology.

        Args:
            products (list): List of product dictionaries
            params (dict): Calculation parameters

        Returns:
            dict: Portfolio ROI metrics
        """
        total_delta_investment = 0
        total_savings = 0
        total_annual_cases = 0
        total_avg_inventory = 0
        total_weighted_days_at_risk = 0
        total_investment_for_weighting = 0

        for product in products:
            result = self.compute_line_item_roi(product, params)

            if 'error' not in result:
                delta_investment = result.get('deltaInvestment', 0)
                savings = result.get('savings', 0)
                days_at_risk = result.get('debug', {}).get('daysAtRisk', 0)
                bulk_cases = product.get('bulk_quantity', 0)

                total_delta_investment += delta_investment
                total_savings += savings

                # Collect data for portfolio deal cycles calculation (raw turnover)
                annual_cases = product.get('annual_cases', 0)
                on_hand_cases = product.get('on_hand', 0)

                total_annual_cases += annual_cases
                total_avg_inventory += (on_hand_cases + bulk_cases) / 2

                # Weight days at risk by actual investment amounts for portfolio calculation
                # This ensures the multiplier reflects true capital exposure
                if delta_investment > 0 and days_at_risk > 0:
                    total_weighted_days_at_risk += days_at_risk * delta_investment
                    total_investment_for_weighting += delta_investment

        # Calculate portfolio ROI
        portfolio_roi = total_savings / total_delta_investment if total_delta_investment > 0 else 0

        # Calculate portfolio deal cycles per year (raw inventory turnover, unadjusted)
        portfolio_deal_cycles = total_annual_cases / total_avg_inventory if total_avg_inventory > 0 else 0

        # Calculate investment-weighted average days at risk for portfolio
        if total_investment_for_weighting > 0:
            weighted_avg_days_at_risk = total_weighted_days_at_risk / total_investment_for_weighting
            # Calculate annual ROI multiplier (after-terms) based on investment-weighted average exposure
            portfolio_roi_multiplier = 365 / weighted_avg_days_at_risk if weighted_avg_days_at_risk > 0 else 0
        else:
            portfolio_roi_multiplier = 0

        return {
            'roi': portfolio_roi,
            'dealCyclesPerYear': portfolio_deal_cycles,
            'annualROIMultiplier': portfolio_roi_multiplier,
            'weightedAvgDaysAtRisk': weighted_avg_days_at_risk if total_investment_for_weighting > 0 else 0
        }

    def calculate(self, data):
        """
        Calculate results for products with the given parameters.

        Args:
            data (dict): Dictionary containing products and parameters

        Returns:
            dict: Calculation results
        """
        try:
            products = data.get('products', [])
            params = data.get('parameters', {})

            # Validate inputs
            if not products:
                raise ValueError("No products provided")

            # Set default parameter values if not provided
            params.setdefault('dealSizeCases', 60)
            params.setdefault('minDaysStock', 30)
            params.setdefault('paymentTermsDays', 30)
            params.setdefault('iterations', 'auto')
            # Handle both old and new parameter names for backward compatibility
            if 'smallDealCases' not in params and 'smallDealMinimum' not in params:
                params['smallDealCases'] = 30

            # Perform initial allocation if bulk_quantity not already set
            for product in products:
                if 'bulk_quantity' not in product:
                    products = self.allocate_based_on_need(products, params['dealSizeCases'], params['minDaysStock'])
                    break

            # Calculate ROI metrics for each product
            results = []
            total_delta_investment = 0
            total_savings = 0

            for product in products:
                metrics = self.compute_line_item_roi(product, params)

                product_copy = product.copy()
                product_copy['metrics'] = metrics

                results.append(product_copy)

                if 'error' not in metrics:
                    total_delta_investment += metrics.get('deltaInvestment', 0)
                    total_savings += metrics.get('savings', 0)

            # Calculate portfolio metrics
            portfolio_metrics = self.calculate_portfolio_roi(products, params)

            return {
                'products': results,
                'totalInvestment': float(total_delta_investment),
                'totalSavings': float(total_savings),
                'portfolioROI': float(portfolio_metrics['roi']),
                'portfolioDealCycles': float(portfolio_metrics['dealCyclesPerYear']),
                'portfolioROIMultiplier': float(portfolio_metrics['annualROIMultiplier']),
                'weightedAvgDaysAtRisk': float(portfolio_metrics['weightedAvgDaysAtRisk'])
            }

        except Exception as e:
            self.logger.error(f"Error calculating results: {str(e)}")
            raise ValueError(f"Calculation error: {str(e)}")

    def optimize(self, data):
        """
        Run optimization on the products.

        Args:
            data (dict): Dictionary containing products and parameters

        Returns:
            dict: Optimization results
        """
        try:
            products = data.get('products', [])
            params = data.get('parameters', {})

            # Validate inputs
            if not products:
                raise ValueError("No products provided")

            # Set default parameter values if not provided
            params.setdefault('dealSizeCases', 60)
            params.setdefault('minDaysStock', 30)
            params.setdefault('paymentTermsDays', 30)
            params.setdefault('iterations', 'auto')

            # Run optimization
            optimization_results = self.run_iterations(products, params)

            # Calculate final metrics
            optimized_data = {
                'products': optimization_results['products'],
                'parameters': params
            }
            calculation_results = self.calculate(optimized_data)

            # Combine results
            return {
                'products': calculation_results['products'],
                'totalInvestment': calculation_results['totalInvestment'],
                'totalSavings': calculation_results['totalSavings'],
                'portfolioROI': calculation_results['portfolioROI'],
                'portfolioDealCycles': calculation_results['portfolioDealCycles'],
                'portfolioROIMultiplier': calculation_results['portfolioROIMultiplier'],
                'weightedAvgDaysAtRisk': calculation_results['weightedAvgDaysAtRisk'],
                'history': optimization_results['history'],
                'totalIterations': optimization_results['totalIterations']
            }

        except Exception as e:
            self.logger.error(f"Error in optimization: {str(e)}")
            raise ValueError(f"Optimization error: {str(e)}")

    def save_scenario(self, data):
        """
        Save a scenario.

        Args:
            data (dict): Scenario data

        Returns:
            bool: True if saved successfully
        """
        try:
            name = data.get('name')
            if not name:
                raise ValueError("Scenario name is required")

            # Ensure the scenarios directory exists
            os.makedirs(self.scenarios_dir, exist_ok=True)

            # Save the scenario
            scenario_path = os.path.join(self.scenarios_dir, f"{name}.json")
            with open(scenario_path, 'w') as f:
                json.dump(data, f, indent=2)

            return True

        except Exception as e:
            self.logger.error(f"Error saving scenario: {str(e)}")
            raise ValueError(f"Error saving scenario: {str(e)}")

    def load_scenario(self, name):
        """
        Load a scenario.

        Args:
            name (str): Scenario name

        Returns:
            dict: Scenario data
        """
        try:
            # Ensure the scenarios directory exists
            os.makedirs(self.scenarios_dir, exist_ok=True)

            # Load the scenario
            scenario_path = os.path.join(self.scenarios_dir, f"{name}.json")
            if not os.path.exists(scenario_path):
                raise ValueError(f"Scenario '{name}' not found")

            with open(scenario_path, 'r') as f:
                return json.load(f)

        except Exception as e:
            self.logger.error(f"Error loading scenario: {str(e)}")
            raise ValueError(f"Error loading scenario: {str(e)}")

    def list_scenarios(self):
        """
        List all available scenarios.

        Returns:
            list: List of scenario names
        """
        try:
            # Ensure the scenarios directory exists
            os.makedirs(self.scenarios_dir, exist_ok=True)

            # List scenario files
            scenario_files = [f for f in os.listdir(self.scenarios_dir) if f.endswith('.json')]
            scenario_names = [os.path.splitext(f)[0] for f in scenario_files]

            return scenario_names

        except Exception as e:
            self.logger.error(f"Error listing scenarios: {str(e)}")
            raise ValueError(f"Error listing scenarios: {str(e)}")

    def delete_scenario(self, name):
        """
        Delete a scenario.

        Args:
            name (str): Scenario name

        Returns:
            bool: True if deleted successfully
        """
        try:
            # Ensure the scenarios directory exists
            os.makedirs(self.scenarios_dir, exist_ok=True)

            # Delete the scenario
            scenario_path = os.path.join(self.scenarios_dir, f"{name}.json")
            if not os.path.exists(scenario_path):
                raise ValueError(f"Scenario '{name}' not found")

            os.remove(scenario_path)
            return True

        except Exception as e:
            self.logger.error(f"Error deleting scenario: {str(e)}")
            raise ValueError(f"Error deleting scenario: {str(e)}")

    def generate_excel_report(self, data):
        """
        Generate an Excel report for the given data.

        Args:
            data (dict): Report data

        Returns:
            str: Path to the generated Excel file
        """
        try:
            name = data.get('name', 'Unnamed')
            products = data.get('products', [])
            params = data.get('parameters', {})
            results = data.get('results', {})
            history = data.get('history', [])

            # Create Excel workbook
            wb = Workbook()

            # Create summary sheet
            summary_sheet = wb.active
            summary_sheet.title = "Summary"

            # Add title
            title = f"Multi-Product Buying Calculator - {name}"
            summary_sheet['A1'] = title
            summary_sheet['A1'].font = get_title_font()
            summary_sheet.merge_cells('A1:F1')

            # Add parameters
            summary_sheet['A3'] = "Parameters"
            summary_sheet['A3'].font = get_header_font()

            summary_sheet['A4'] = "Deal Size:"
            summary_sheet['B4'] = params.get('dealSizeCases', 60)
            summary_sheet['C4'] = "cases"

            summary_sheet['A5'] = "Minimum Days Stock:"
            summary_sheet['B5'] = params.get('minDaysStock', 30)
            summary_sheet['C5'] = "days"

            summary_sheet['A6'] = "Payment Terms:"
            summary_sheet['B6'] = params.get('paymentTermsDays', 30)
            summary_sheet['C6'] = "days"

            summary_sheet['A7'] = "Optimization Iterations:"
            summary_sheet['B7'] = params.get('iterations', 'auto')

            # Add summary results
            summary_sheet['A9'] = "Summary Results"
            summary_sheet['A9'].font = get_header_font()

            summary_sheet['A10'] = "Total Products:"
            summary_sheet['B10'] = len(products)

            summary_sheet['A11'] = "Total Additional Investment:"
            summary_sheet['B11'] = results.get('totalInvestment', 0)
            summary_sheet['B11'].number_format = get_money_format()

            summary_sheet['A12'] = "Total Savings:"
            summary_sheet['B12'] = results.get('totalSavings', 0)
            summary_sheet['B12'].number_format = get_money_format()

            summary_sheet['A13'] = "ROI:"
            summary_sheet['B13'] = results.get('portfolioROI', 0)
            summary_sheet['B13'].number_format = get_percent_format()

            summary_sheet['A14'] = "Raw Inventory Turnover (Unadjusted):"
            summary_sheet['B14'] = results.get('portfolioDealCycles', 0)

            summary_sheet['A15'] = "Annual ROI Multiplier (After-Terms):"
            summary_sheet['B15'] = results.get('portfolioROIMultiplier', 0)

            summary_sheet['A16'] = "Annualized ROI:"
            annualized_roi = results.get('portfolioROI', 0) * results.get('portfolioROIMultiplier', 0)
            summary_sheet['B16'] = annualized_roi
            summary_sheet['B16'].number_format = get_percent_format()

            # Create detailed results sheet
            details_sheet = wb.create_sheet(title="Detailed Results")

            # Add headers
            headers = ["Product", "Price (Small)", "Price (Bulk)", "Cases On Hand",
                      "Annual Cases", "Bottles/Case", "Bulk Cases", "Days of Stock",
                      "Savings", "Investment", "ROI", "Deal Cycles", "Annual ROI"]

            for col, header in enumerate(headers, 1):
                cell = details_sheet.cell(row=1, column=col, value=header)
                cell.font = get_header_font()
                cell.fill = get_header_fill()

            # Add product data
            for row, product in enumerate(products, 2):
                metrics = product.get('metrics', {})

                details_sheet.cell(row=row, column=1, value=product.get('name', ''))
                details_sheet.cell(row=row, column=2, value=product.get('priceSmall', 0)).number_format = get_money_format()
                details_sheet.cell(row=row, column=3, value=product.get('priceBulk', 0)).number_format = get_money_format()
                details_sheet.cell(row=row, column=4, value=product.get('onHandCases', 0))
                details_sheet.cell(row=row, column=5, value=product.get('annualCases', 0))
                details_sheet.cell(row=row, column=6, value=product.get('bottlesPerCase', 0))
                details_sheet.cell(row=row, column=7, value=product.get('bulkCases', 0))

                # Skip metrics if there's an error
                if 'error' in metrics:
                    error_msg = f"Error: {metrics['error']}"
                    for col in range(8, 13):
                        details_sheet.cell(row=row, column=col, value=error_msg)
                else:
                    details_sheet.cell(row=row, column=8, value=product.get('daysOfStock', 0))
                    details_sheet.cell(row=row, column=9, value=metrics.get('savings', 0)).number_format = get_money_format()
                    details_sheet.cell(row=row, column=10, value=metrics.get('deltaInvestment', 0)).number_format = get_money_format()
                    details_sheet.cell(row=row, column=11, value=metrics.get('roi', 0)).number_format = get_percent_format()
                    details_sheet.cell(row=row, column=12, value=metrics.get('dealCyclesPerYear', 0))
                    details_sheet.cell(row=row, column=13, value=metrics.get('annualROIMultiplier', 0))

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                details_sheet.column_dimensions[column_letter].width = 15

            # Create optimization history sheet if history exists
            if history:
                history_sheet = wb.create_sheet(title="Optimization History")

                # Add headers
                history_headers = ["Iteration", "Action", "Portfolio ROI", "Change"]

                for col, header in enumerate(history_headers, 1):
                    cell = history_sheet.cell(row=1, column=col, value=header)
                    cell.font = get_header_font()
                    cell.fill = get_header_fill()

                # Add history data
                prev_roi = 0
                for row, iter_data in enumerate(history, 2):
                    iteration = iter_data.get('iteration', 0)
                    roi = iter_data.get('totalROI', 0)

                    # Action description
                    if iteration == 0:
                        action = "Initial Allocation"
                    else:
                        swapped = iter_data.get('swapped', {})
                        from_product = swapped.get('from', '')
                        to_product = swapped.get('to', '')
                        action = f"Moved 1 case from {from_product} to {to_product}"

                    # ROI change
                    roi_change = roi - prev_roi if row > 2 else 0

                    history_sheet.cell(row=row, column=1, value=iteration)
                    history_sheet.cell(row=row, column=2, value=action)
                    history_sheet.cell(row=row, column=3, value=roi).number_format = get_percent_format()
                    history_sheet.cell(row=row, column=4, value=roi_change).number_format = get_percent_format()

                    prev_roi = roi

                # Create ROI chart
                chart_sheet = wb.create_sheet(title="ROI Chart")

                # Add chart title
                chart_sheet['A1'] = "ROI Optimization Progress"
                chart_sheet['A1'].font = get_title_font()

                # Add data for chart
                chart_sheet['A3'] = "Iteration"
                chart_sheet['B3'] = "Portfolio ROI"

                for row, iter_data in enumerate(history, 4):
                    chart_sheet.cell(row=row, column=1, value=iter_data.get('iteration', 0))
                    chart_sheet.cell(row=row, column=2, value=iter_data.get('totalROI', 0))

                # Create chart
                chart = LineChart()
                chart.title = "ROI Optimization Progress"
                chart.style = 2
                chart.x_axis.title = "Iteration"
                chart.y_axis.title = "Portfolio ROI"

                data = Reference(chart_sheet, min_col=2, min_row=3, max_row=len(history) + 3, max_col=2)
                cats = Reference(chart_sheet, min_col=1, min_row=4, max_row=len(history) + 3)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

                chart_sheet.add_chart(chart, "D4")

            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"multi_product_report_{name.replace(' ', '_')}_{timestamp}.xlsx"
            report_dir = "reports"
            os.makedirs(report_dir, exist_ok=True)
            filepath = os.path.join(report_dir, filename)

            # Save workbook
            wb.save(filepath)

            return filepath

        except Exception as e:
            self.logger.error(f"Error generating Excel report: {str(e)}")
            raise ValueError(f"Error generating Excel report: {str(e)}")

    def generate_distributor_order_excel(self, data):
        """
        Generate a clean Excel order sheet for distributors.

        Args:
            data (dict): Order data containing products and parameters

        Returns:
            str: Path to the generated Excel file
        """
        try:
            name = data.get('name', 'Order')
            products = data.get('products', [])
            params = data.get('parameters', {})

            # Create Excel workbook
            wb = Workbook()
            order_sheet = wb.active
            order_sheet.title = "Distributor Order"

            # Add title and order info
            order_sheet['A1'] = "PURCHASE ORDER"
            order_sheet['A1'].font = Font(size=18, bold=True)
            order_sheet.merge_cells('A1:E1')

            # Add order details
            order_sheet['A3'] = f"Order Name: {name}"
            order_sheet['A3'].font = Font(size=12, bold=True)

            order_sheet['A4'] = f"Date: {datetime.now().strftime('%B %d, %Y')}"
            order_sheet['A5'] = f"Total Cases: {params.get('dealSizeCases', 0)} cases"

            # Add product table headers
            headers = ["Product Name", "Bulk Price", "Cases to Order", "Bottles per Case", "Total Bottles", "Line Total"]
            header_row = 7

            for col, header in enumerate(headers, 1):
                cell = order_sheet.cell(row=header_row, column=col, value=header)
                cell.font = get_header_font()
                cell.fill = get_header_fill()
                cell.alignment = Alignment(horizontal='center')

            # Add product data
            total_cases = 0
            total_bottles = 0
            total_cost = 0.0

            for row, product in enumerate(products, header_row + 1):
                if product.get('bulkCases', 0) > 0:  # Only include products with bulk cases
                    bulk_cases = product.get('bulkCases', 0)
                    bulk_price = product.get('priceBulk', 0)
                    bottles_per_case = product.get('bottlesPerCase', 12)
                    total_bottles_product = bulk_cases * bottles_per_case
                    line_total = bulk_cases * bulk_price

                    order_sheet.cell(row=row, column=1, value=product.get('name', ''))
                    order_sheet.cell(row=row, column=2, value=bulk_price).number_format = '$#,##0.00'
                    order_sheet.cell(row=row, column=3, value=bulk_cases)
                    order_sheet.cell(row=row, column=4, value=bottles_per_case)
                    order_sheet.cell(row=row, column=5, value=total_bottles_product)
                    order_sheet.cell(row=row, column=6, value=line_total).number_format = '$#,##0.00'

                    total_cases += bulk_cases
                    total_bottles += total_bottles_product
                    total_cost += line_total

            # Add totals row
            last_row = header_row + len([p for p in products if p.get('bulkCases', 0) > 0]) + 1

            order_sheet.cell(row=last_row, column=1, value="TOTALS").font = Font(bold=True)
            order_sheet.cell(row=last_row, column=2, value="").font = Font(bold=True)
            order_sheet.cell(row=last_row, column=3, value=total_cases).font = Font(bold=True)
            order_sheet.cell(row=last_row, column=4, value="")
            order_sheet.cell(row=last_row, column=5, value=total_bottles).font = Font(bold=True)
            order_sheet.cell(row=last_row, column=6, value=total_cost).font = Font(bold=True)
            order_sheet.cell(row=last_row, column=6).number_format = '$#,##0.00'

            # Style the totals row
            for col in range(1, 7):
                cell = order_sheet.cell(row=last_row, column=col)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick')
                )

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                order_sheet.column_dimensions[column_letter].width = 15

            # Make product name column wider
            order_sheet.column_dimensions['A'].width = 25

            # Add notes section
            notes_row = last_row + 3
            order_sheet.cell(row=notes_row, column=1, value="Notes:").font = Font(bold=True)
            order_sheet.cell(row=notes_row + 1, column=1, value="• This order is optimized for bulk pricing")
            order_sheet.cell(row=notes_row + 2, column=1, value="• Please confirm availability before processing")
            order_sheet.cell(row=notes_row + 3, column=1, value="• Contact us with any questions about quantities")

            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = name.replace(' ', '_').replace('/', '_').replace('\\', '_')
            filename = f"distributor_order_{safe_name}_{timestamp}.xlsx"
            report_dir = "reports"
            os.makedirs(report_dir, exist_ok=True)
            filepath = os.path.join(report_dir, filename)

            # Save workbook
            wb.save(filepath)

            self.logger.info(f"Generated distributor order Excel: {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(f"Error generating distributor order Excel: {str(e)}")
            raise ValueError(f"Error generating distributor order Excel: {str(e)}")

# For backward compatibility
if __name__ == "__main__":
    calculator = MultiProductBuyingCalculator()
    print("Multi-Product Buying Calculator initialized.")
