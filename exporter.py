"""
Export module for the multi-product calculator.
Provides functions for exporting scenarios to Excel and debugging formats.
"""

import os
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart

class ExportError(Exception):
    """Custom exception for export errors."""
    pass

class MultiProductExporter:
    """Exports multi-product calculator data to various formats."""

    def __init__(self, output_dir="reports"):
        """
        Initialize the exporter.

        Parameters:
        - output_dir: Directory to store output files
        """
        self.output_dir = output_dir
        self._ensure_output_dir()

    def _ensure_output_dir(self):
        """Ensure the output directory exists."""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def export_to_excel(self, scenario_name, scenario_data, product_metrics):
        """
        Export scenario data to Excel.

        Parameters:
        - scenario_name: Name of the scenario
        - scenario_data: Dictionary with scenario data
        - product_metrics: Dictionary with product metrics

        Returns: Path to the output file
        Raises ExportError if export fails
        """
        try:
            # Ensure output directory exists
            self._ensure_output_dir()

            # Create workbook
            wb = openpyxl.Workbook()

            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # Create sheets
            summary_sheet = wb.create_sheet("Summary")
            products_sheet = wb.create_sheet("Products")
            details_sheet = wb.create_sheet("Details")

            # Styles
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            title_font = Font(size=14, bold=True)
            header_font = Font(bold=True)
            money_format = '#,##0.00_);[Red](#,##0.00)'
            percent_format = '0.00%;[Red](0.00%)'

            # Create summary sheet
            self._create_summary_sheet(summary_sheet, scenario_data, product_metrics, title_font, header_fill, money_format, percent_format)

            # Create products sheet
            self._create_products_sheet(products_sheet, scenario_data["products"], product_metrics, title_font, header_fill, money_format, percent_format)

            # Create details sheet
            self._create_details_sheet(details_sheet, scenario_data, product_metrics, title_font, header_fill, money_format, percent_format)

            # Auto-size columns
            for sheet in wb.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            file_path = os.path.join(self.output_dir, f"{scenario_name}.xlsx")
            wb.save(file_path)

            return file_path

        except Exception as e:
            raise ExportError(f"Error exporting to Excel: {e}")

    def export_debug_info(self, scenario_name, scenario_data, product_metrics):
        """
        Export debug information to JSON.

        Parameters:
        - scenario_name: Name of the scenario
        - scenario_data: Dictionary with scenario data
        - product_metrics: Dictionary with product metrics

        Returns: Path to the output file
        Raises ExportError if export fails
        """
        try:
            # Ensure output directory exists
            self._ensure_output_dir()

            # Create debug info
            debug_info = {
                "scenario_name": scenario_name,
                "input_data": scenario_data,
                "calculation_results": product_metrics
            }

            # Sanitize data for JSON serialization
            sanitized_data = self._sanitize_for_json(debug_info)

            # Save to file
            file_path = os.path.join(self.output_dir, f"{scenario_name}_debug.json")
            with open(file_path, 'w') as f:
                json.dump(sanitized_data, f, indent=2)

            return file_path

        except Exception as e:
            raise ExportError(f"Error exporting debug info: {e}")

    def _create_summary_sheet(self, sheet, scenario_data, product_metrics, title_font, header_fill, money_format, percent_format):
        """Create the summary sheet."""
        # Title
        sheet.cell(row=1, column=1, value="Multi-Product Deal Calculator - Summary")
        sheet.cell(row=1, column=1).font = title_font

        # Parameters
        sheet.cell(row=3, column=1, value="Parameters")
        sheet.cell(row=3, column=1).font = Font(bold=True)

        params = scenario_data["params"]
        sheet.cell(row=4, column=1, value="Small Deal Minimum")
        sheet.cell(row=4, column=2, value=params["small_deal_minimum"])

        sheet.cell(row=5, column=1, value="Bulk Deal Minimum")
        sheet.cell(row=5, column=2, value=params["bulk_deal_minimum"])

        sheet.cell(row=6, column=1, value="Payment Terms")
        sheet.cell(row=6, column=2, value=params["payment_terms"])
        sheet.cell(row=6, column=3, value="days")

        # Summary Results
        sheet.cell(row=8, column=1, value="Summary Results")
        sheet.cell(row=8, column=1).font = Font(bold=True)

        sheet.cell(row=9, column=1, value="Peak Additional Investment")
        sheet.cell(row=9, column=2, value=product_metrics["total_peak_investment"])
        sheet.cell(row=9, column=2).number_format = money_format

        sheet.cell(row=10, column=1, value="Average Investment")
        sheet.cell(row=10, column=2, value=product_metrics["total_average_investment"])
        sheet.cell(row=10, column=2).number_format = money_format

        sheet.cell(row=11, column=1, value="Total Savings")
        sheet.cell(row=11, column=2, value=product_metrics["total_savings"])
        sheet.cell(row=11, column=2).number_format = money_format

        sheet.cell(row=12, column=1, value="ROI")
        sheet.cell(row=12, column=2, value=product_metrics["overall_roi"])
        sheet.cell(row=12, column=2).number_format = percent_format

        sheet.cell(row=13, column=1, value="Annualized ROI")
        sheet.cell(row=13, column=2, value=product_metrics["overall_annual_roi"])
        sheet.cell(row=13, column=2).number_format = percent_format

    def _create_products_sheet(self, sheet, products, product_metrics, title_font, header_fill, money_format, percent_format):
        """Create the products sheet."""
        # Title
        sheet.cell(row=1, column=1, value="Multi-Product Deal Calculator - Products")
        sheet.cell(row=1, column=1).font = title_font

        # Headers
        headers = [
            "Product", "Current Price", "Bulk Price", "Cases On Hand",
            "Cases/Year", "Bottles/Case", "Bulk Quantity", "Total Inventory",
            "Days of Stock", "ROI"
        ]

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Product data
        product_metrics_dict = {pm["product_name"]: pm for pm in product_metrics["product_metrics"]}

        for row, product in enumerate(products, start=4):
            product_name = product["product_name"]
            pm = product_metrics_dict.get(product_name, {})

            sheet.cell(row=row, column=1, value=product_name)
            sheet.cell(row=row, column=2, value=product["current_price"])
            sheet.cell(row=row, column=2).number_format = money_format

            sheet.cell(row=row, column=3, value=product["bulk_price"])
            sheet.cell(row=row, column=3).number_format = money_format

            sheet.cell(row=row, column=4, value=product["cases_on_hand"])

            sheet.cell(row=row, column=5, value=product["cases_per_year"])

            sheet.cell(row=row, column=6, value=product["bottles_per_case"])

            sheet.cell(row=row, column=7, value=product.get("bulk_quantity", 0))

            total_inventory = product["cases_on_hand"] + product.get("bulk_quantity", 0)
            sheet.cell(row=row, column=8, value=total_inventory)

            sheet.cell(row=row, column=9, value=pm.get("days_of_stock_after", 0))

            sheet.cell(row=row, column=10, value=pm.get("roi", 0))
            sheet.cell(row=row, column=10).number_format = percent_format

        # Total row
        total_row = len(products) + 4
        sheet.cell(row=total_row, column=1, value="Total")
        sheet.cell(row=total_row, column=1).font = Font(bold=True)

        # Sum columns that should be totaled
        columns_to_total = [7]  # Bulk Quantity
        for col in columns_to_total:
            sheet.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})")

    def _create_details_sheet(self, sheet, scenario_data, product_metrics, title_font, header_fill, money_format, percent_format):
        """Create the details sheet."""
        # Title
        sheet.cell(row=1, column=1, value="Multi-Product Deal Calculator - Details")
        sheet.cell(row=1, column=1).font = title_font

        # Headers
        headers = [
            "Product", "Current Price", "Bulk Price", "Bulk Quantity",
            "Savings Per Case", "Total Savings", "ROI"
        ]

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Product data
        product_metrics_dict = {pm["product_name"]: pm for pm in product_metrics["product_metrics"]}

        for row, product in enumerate(scenario_data["products"], start=4):
            product_name = product["product_name"]
            pm = product_metrics_dict.get(product_name, {})

            sheet.cell(row=row, column=1, value=product_name)

            sheet.cell(row=row, column=2, value=product["current_price"])
            sheet.cell(row=row, column=2).number_format = money_format

            sheet.cell(row=row, column=3, value=product["bulk_price"])
            sheet.cell(row=row, column=3).number_format = money_format

            sheet.cell(row=row, column=4, value=product.get("bulk_quantity", 0))

            sheet.cell(row=row, column=5, value=pm.get("savings_per_case", 0))
            sheet.cell(row=row, column=5).number_format = money_format

            sheet.cell(row=row, column=6, value=pm.get("total_savings", 0))
            sheet.cell(row=row, column=6).number_format = money_format

            sheet.cell(row=row, column=7, value=pm.get("roi", 0))
            sheet.cell(row=row, column=7).number_format = percent_format

        # Total row
        total_row = len(scenario_data["products"]) + 4
        sheet.cell(row=total_row, column=1, value="Total")
        sheet.cell(row=total_row, column=1).font = Font(bold=True)

        # Sum columns that should be totaled
        columns_to_total = [4, 6]  # Bulk Quantity, Total Savings
        for col in columns_to_total:
            sheet.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})")

        # Overall ROI
        sheet.cell(row=total_row, column=7, value=product_metrics["overall_roi"])
        sheet.cell(row=total_row, column=7).number_format = percent_format

    def _sanitize_for_json(self, data):
        """
        Sanitize data for JSON serialization.

        Parameters:
        - data: Data to sanitize

        Returns: Sanitized data
        """
        if isinstance(data, dict):
            return {k: self._sanitize_for_json(v) for k, v in data.items()}

        elif isinstance(data, list):
            return [self._sanitize_for_json(item) for item in data]

        elif hasattr(data, 'item'):  # numpy types have .item() method
            try:
                return data.item()
            except:
                return data

        elif isinstance(data, (int, float, str, bool, type(None))):
            return data

        else:
            return str(data)  # Convert any other types to string
