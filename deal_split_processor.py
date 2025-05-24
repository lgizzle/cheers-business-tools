import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
import os
from datetime import datetime
from scenario_utils import load_all_scenarios, save_all_scenarios
from logging_utils import setup_logging
from excel_utils import get_title_font, get_header_font, get_header_fill, get_money_format, get_percent_format, apply_header_styles

setup_logging('logs/deal_split.log')

class DealSplitCalculator:
    def __init__(self):
        self.scenarios = {}
        self.scenarios_file = os.path.join(os.path.dirname(__file__), 'scenarios.json')
        self.load_scenarios()

    def load_scenarios(self):
        """Load existing scenarios from a JSON file if it exists"""
        self.scenarios = load_all_scenarios(self.scenarios_file)

    def save_scenarios(self):
        """Save scenarios to a JSON file"""
        save_all_scenarios(self.scenarios_file, self.scenarios)

    def calculate_split(self, data, desired_total):
        """
        Calculate deal splits based on annual sales proportions

        Args:
            data: List of dictionaries with 'variety', 'annual_sales', and 'inventory_on_hand' keys
            desired_total: The desired total order quantity

        Returns:
            DataFrame with calculation results
        """
        # Create DataFrame from input data
        df = pd.DataFrame(data)

        # Ensure inventory_on_hand exists, default to 0 if missing
        if 'inventory_on_hand' not in df.columns:
            df['inventory_on_hand'] = 0

        # Calculate totals
        total_annual_sales = df['annual_sales'].sum()

        # Calculate proportions and splits
        if total_annual_sales > 0:
            df['calculated_split'] = df['annual_sales'] * desired_total / total_annual_sales
            df['rounded_split'] = df['calculated_split'].apply(lambda x: int(x))

            # Calculate days to sell for each variety
            # Daily sales rate = annual sales / 365
            # Days to sell = rounded split / daily sales rate
            df['daily_sales_rate'] = df['annual_sales'] / 365
            df['days_to_sell'] = df.apply(
                lambda row: round(row['rounded_split'] / row['daily_sales_rate']) if row['daily_sales_rate'] > 0 else 0,
                axis=1
            )

            # If the sum of rounded_split doesn't match the desired total,
            # adjust values with the largest difference between calculated and rounded
            rounded_total = df['rounded_split'].sum()

            if rounded_total != desired_total:
                # Calculate difference
                df['diff'] = df['calculated_split'] - df['rounded_split']

                # Sort by difference
                df = df.sort_values('diff', ascending=False)

                # Adjust values
                difference = desired_total - rounded_total

                if difference > 0:
                    # Need to add units
                    for i in range(int(difference)):
                        if i < len(df):
                            df.iloc[i, df.columns.get_loc('rounded_split')] += 1

                            # Recalculate days to sell for this row
                            daily_rate = df.iloc[i]['daily_sales_rate']
                            if daily_rate > 0:
                                df.iloc[i, df.columns.get_loc('days_to_sell')] = round(df.iloc[i]['rounded_split'] / daily_rate)
                else:
                    # Need to remove units
                    for i in range(int(abs(difference))):
                        if i < len(df):
                            idx = len(df) - 1 - i  # Start from bottom
                            if df.iloc[idx]['rounded_split'] > 0:  # Don't go below zero
                                df.iloc[idx, df.columns.get_loc('rounded_split')] -= 1

                                # Recalculate days to sell for this row
                                daily_rate = df.iloc[idx]['daily_sales_rate']
                                if daily_rate > 0:
                                    df.iloc[idx, df.columns.get_loc('days_to_sell')] = round(df.iloc[idx]['rounded_split'] / daily_rate)

                # Reset index and sort
                df = df.sort_index()

                # Remove diff column
                df = df.drop('diff', axis=1)
        else:
            # Handle case where there are no annual sales
            df['calculated_split'] = 0
            df['rounded_split'] = 0
            df['daily_sales_rate'] = 0
            df['days_to_sell'] = 0

        # Convert NumPy types to Python native types for JSON serialization
        df['annual_sales'] = df['annual_sales'].astype(int)
        df['inventory_on_hand'] = df['inventory_on_hand'].astype(int)
        df['calculated_split'] = df['calculated_split'].astype(float)
        df['rounded_split'] = df['rounded_split'].astype(int)
        df['days_to_sell'] = df['days_to_sell'].astype(int)

        # Drop the daily_sales_rate column as it's only needed for calculation
        if 'daily_sales_rate' in df.columns:
            df = df.drop('daily_sales_rate', axis=1)

        return df

    def save_scenario(self, name, data):
        """
        Save a scenario for future use

        Args:
            name: Name of the scenario
            data: List of dictionaries with variety and annual_sales
        """
        self.scenarios[name] = data
        self.save_scenarios()

    def get_scenario(self, name):
        """
        Retrieve a saved scenario

        Args:
            name: Name of the scenario

        Returns:
            List of dictionaries or None if scenario doesn't exist
        """
        return self.scenarios.get(name)

    def delete_scenario(self, name):
        """Delete a scenario by name"""
        if name in self.scenarios:
            del self.scenarios[name]
            self.save_scenarios()
            return True
        return False

    def get_all_scenarios(self):
        """Return all saved scenarios"""
        return self.scenarios

    def generate_report(self, data, desired_total, file_path):
        """
        Generate an Excel report with the deal split calculation

        Args:
            data: List of dictionaries with 'variety', 'annual_sales', and 'inventory_on_hand' keys
            desired_total: The desired total order quantity
            file_path: Path to save the Excel file

        Returns:
            Path to the generated Excel file
        """
        # Calculate splits
        results_df = self.calculate_split(data, desired_total)

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deal Calculator"

        # Add title
        ws['A1'] = "DEAL SPLIT CALCULATOR"
        ws['A1'].font = get_title_font()
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Add headers
        headers = ['Variety', 'Annual Sales', 'Inventory on Hand', 'Calculated Split', 'Rounded Split', 'Actual Order', 'Days to Sell']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
        apply_header_styles(ws, row=3, start_col=1, end_col=len(headers))

        # Add data
        row_num = 4
        for _, row in results_df.iterrows():
            ws.cell(row=row_num, column=1).value = row['variety']
            ws.cell(row=row_num, column=2).value = row['annual_sales']
            ws.cell(row=row_num, column=3).value = row['inventory_on_hand']
            ws.cell(row=row_num, column=4).value = row['calculated_split']
            ws.cell(row=row_num, column=4).number_format = get_money_format()
            ws.cell(row=row_num, column=5).value = row['rounded_split']
            ws.cell(row=row_num, column=6).value = row['rounded_split']  # Default actual order to rounded
            ws.cell(row=row_num, column=7).value = row['days_to_sell']
            row_num += 1

        # Add totals row
        total_row = row_num
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=1).font = get_header_font()

        # Sum formulas
        ws.cell(row=total_row, column=2).value = f"=SUM(B4:B{total_row-1})"
        ws.cell(row=total_row, column=3).value = f"=SUM(C4:C{total_row-1})"
        ws.cell(row=total_row, column=4).value = f"=F{total_row+2}"  # Link to desired total
        ws.cell(row=total_row, column=5).value = f"=SUM(E4:E{total_row-1})"
        ws.cell(row=total_row, column=6).value = f"=SUM(F4:F{total_row-1})"
        # Average days to sell (weighted by order quantity)
        ws.cell(row=total_row, column=7).value = f"=SUMPRODUCT(F4:F{total_row-1},G4:G{total_row-1})/F{total_row}"
        ws.cell(row=total_row, column=7).number_format = '0'

        # Add desired total
        ws.cell(row=total_row+2, column=1).value = "Desired Total Order:"
        ws.cell(row=total_row+2, column=1).font = get_header_font()
        ws.cell(row=total_row+2, column=6).value = desired_total

        # Add status message
        ws.cell(row=total_row+4, column=1).value = "Status:"
        ws.cell(row=total_row+4, column=1).font = get_header_font()
        status_formula = (
            f'=IF(AND(F{total_row+2}=0,SUM(B4:B{total_row-1})=0),"Enter sales data and desired total to begin.",'
            f'IF(F{total_row}=F{total_row+2},"Perfect! Your order total matches your target.",'
            f'IF(F{total_row}<F{total_row+2},"Add "&(F{total_row+2}-F{total_row})&" more units to reach your target.",'
            f'"Remove "&(F{total_row}-F{total_row+2})&" units to reach your target.")))'
        )
        ws.cell(row=total_row+4, column=2).value = status_formula
        ws.merge_cells(f'B{total_row+4}:G{total_row+4}')

        # Set explicit column widths
        column_widths = {
            'A': 25,  # Variety
            'B': 15,  # Annual Sales
            'C': 15,  # Inventory on Hand
            'D': 15,  # Calculated Split
            'E': 15,  # Rounded Split
            'F': 15,  # Actual Order
            'G': 15,  # Days to Sell
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Format number columns
        for row in range(4, total_row + 1):
            ws.cell(row=row, column=2).number_format = '#,##0'    # Annual Sales
            ws.cell(row=row, column=3).number_format = '#,##0'    # Inventory on Hand
            ws.cell(row=row, column=4).number_format = get_money_format()     # Calculated Split
            ws.cell(row=row, column=5).number_format = '#,##0'    # Rounded Split
            ws.cell(row=row, column=6).number_format = '#,##0'    # Actual Order
            ws.cell(row=row, column=7).number_format = '#,##0'    # Days to Sell

        # Save the workbook
        wb.save(file_path)

        return file_path
