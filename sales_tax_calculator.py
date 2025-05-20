import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import json
from datetime import datetime

class SalesTaxCalculator:
    """
    Sales Tax Calculator that works backwards from tax amounts collected to determine
    actual sales amounts and properly distribute taxes to appropriate authorities.
    """

    def __init__(self):
        """Initialize the calculator with tax rates and jurisdiction data."""
        # Tax rates and jurisdictions - matches data from the Constants sheet
        self.jurisdictions = {
            "040206": {"name": "UNINCORPORATED", "city_name": "FALCON", "county": "EL PASO",
                      "city_tax": 0.0000, "county_tax": 0.0123, "rta_tax": 0.0100, "state_tax": 0.0290,
                      "total_tax": 0.0513, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "State Collected"},
            "040206-1": {"name": "UNINCORPORATED", "city_name": "SECURITY/WIDEFIELD", "county": "EL PASO",
                      "city_tax": 0.0000, "county_tax": 0.0123, "rta_tax": 0.0100, "state_tax": 0.0290,
                      "total_tax": 0.0513, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "State Collected"},
            "040206-2": {"name": "UNINCORPORATED", "city_name": "CHIPITA PARK", "county": "EL PASO",
                      "city_tax": 0.0000, "county_tax": 0.0123, "rta_tax": 0.0100, "state_tax": 0.0290,
                      "total_tax": 0.0513, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "State Collected"},
            "040206-3": {"name": "UNINCORPORATED", "city_name": "PEYTON", "county": "EL PASO",
                      "city_tax": 0.0000, "county_tax": 0.0123, "rta_tax": 0.0100, "state_tax": 0.0290,
                      "total_tax": 0.0513, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "State Collected"},
            "040017": {"name": "COLORADO SPRINGS", "city_name": "COLORADO SPRINGS", "county": "EL PASO",
                      "city_tax": 0.0307, "county_tax": 0.0123, "rta_tax": 0.0100, "state_tax": 0.0290,
                      "total_tax": 0.0820, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "Self Collected"},
            "040059": {"name": "PALMER LAKE", "city_name": "PALMER LAKE", "county": "EL PASO",
                      "city_tax": 0.0300, "county_tax": 0.0123, "rta_tax": 0.0000, "state_tax": 0.0290,
                      "total_tax": 0.0713, "state_service_fee": 0.0400, "city_service_fee": 0.0330,
                      "city_collection": "State Collected"},
            "040052": {"name": "MANITOU SPRINGS", "city_name": "MANITOU SPRINGS", "county": "EL PASO",
                      "city_tax": 0.0390, "county_tax": 0.0123, "rta_tax": 0.0100, "state_tax": 0.0290,
                      "total_tax": 0.0903, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "State Collected"},
            "040097": {"name": "WOODLAND PARK", "city_name": "WOODLAND PARK", "county": "EL PASO",
                      "city_tax": 0.0300, "county_tax": 0.0123, "rta_tax": 0.0100, "state_tax": 0.0290,
                      "total_tax": 0.0813, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "Self Collected"},
            "040031": {"name": "FOUNTAIN", "city_name": "FOUNTAIN", "county": "EL PASO",
                      "city_tax": 0.0340, "county_tax": 0.0123, "rta_tax": 0.0000, "state_tax": 0.0290,
                      "total_tax": 0.0753, "state_service_fee": 0.0400, "city_service_fee": 0.0000,
                      "city_collection": "State Collected"},
            "040057": {"name": "MONUMENT", "city_name": "MONUMENT", "county": "EL PASO",
                      "city_tax": 0.0350, "county_tax": 0.0123, "rta_tax": 0.0000, "state_tax": 0.0290,
                      "total_tax": 0.0763, "state_service_fee": 0.0400, "city_service_fee": 0.0330,
                      "city_collection": "Self Collected"},
        }

    def calculate_sales_from_tax(self, tax_data):
        """
        Calculate sales amounts and tax distribution based on collected tax amounts

        Args:
            tax_data: List of dictionaries containing jurisdiction codes and tax amounts
                Each dictionary should have:
                - jurisdiction_code: The jurisdiction code
                - city_name: The city name
                - standard_tax: The standard tax amount collected
                - cigarette_tax: Optional cigarette tax amount collected
                - soda_tax: Optional soda tax amount collected

        Returns:
            Dict containing calculation results
        """
        # Initialize result DataFrames
        sales_results = []
        tax_calcs = []

        # Process each jurisdiction
        total_tax_collected = 0
        total_standard_sales = 0
        total_cig_sales = 0
        total_soda_sales = 0
        total_sales = 0

        for item in tax_data:
            code = item['jurisdiction_code']

            # Handle unincorporated areas with same code but different names
            if code == '040206':
                city_name = item['city_name']
                # Find the correct entry by matching city name
                for key, data in self.jurisdictions.items():
                    if key.startswith('040206') and data['city_name'] == city_name:
                        code = key
                        break
                # If no match found, use default 040206
                if code == '040206' and city_name != 'FALCON':
                    code = '040206'  # Default unincorporated

            if code not in self.jurisdictions:
                # Skip invalid jurisdiction codes
                continue

            # Get jurisdiction tax rates
            jurisdiction = self.jurisdictions[code]

            # Get tax amounts collected or default to zero
            standard_tax = float(item.get('standard_tax', 0) or 0)
            cigarette_tax = float(item.get('cigarette_tax', 0) or 0)
            soda_tax = float(item.get('soda_tax', 0) or 0)

            # Calculate total tax collected for this jurisdiction
            total_jurisdiction_tax = standard_tax + cigarette_tax + soda_tax

            # Calculate sales amount for each tax category by working backwards from tax collected
            # Sales = Tax / Tax Rate
            total_tax_rate = jurisdiction['total_tax']
            if total_tax_rate > 0:
                standard_sales = standard_tax / total_tax_rate if standard_tax > 0 else 0
                cigarette_sales = cigarette_tax / total_tax_rate if cigarette_tax > 0 else 0
                soda_sales = soda_tax / total_tax_rate if soda_tax > 0 else 0
                total_jurisdiction_sales = standard_sales + cigarette_sales + soda_sales
            else:
                standard_sales = 0
                cigarette_sales = 0
                soda_sales = 0
                total_jurisdiction_sales = 0

            # Calculate individual tax components
            city_tax_amount = jurisdiction['city_tax'] * total_jurisdiction_sales
            county_tax_amount = jurisdiction['county_tax'] * total_jurisdiction_sales
            rta_tax_amount = jurisdiction['rta_tax'] * total_jurisdiction_sales
            state_tax_amount = jurisdiction['state_tax'] * total_jurisdiction_sales
            calculated_total_tax = city_tax_amount + county_tax_amount + rta_tax_amount + state_tax_amount

            # Calculate service fees
            city_service_fee = city_tax_amount * jurisdiction['city_service_fee']
            state_service_fee = state_tax_amount * jurisdiction['state_service_fee']

            # Calculate net tax amounts (after service fees)
            state_net_tax = state_tax_amount + county_tax_amount + rta_tax_amount - state_service_fee

            # Determine if city tax goes to state or directly to city
            city_net_tax = 0
            if jurisdiction['city_collection'] == 'Self Collected':
                city_net_tax = city_tax_amount

            # Total net tax
            total_net_tax = state_net_tax + city_net_tax

            # Add to sales results
            sales_results.append({
                'jurisdiction_code': item['jurisdiction_code'],
                'city_name': item['city_name'],
                'jurisdiction_name': jurisdiction['name'],
                'total_tax_collected': total_jurisdiction_tax,
                'standard_sales': standard_sales,
                'cigarette_sales': cigarette_sales,
                'soda_sales': soda_sales,
                'total_sales': total_jurisdiction_sales,
                'calculated_state_net_tax': state_net_tax,
                'calculated_city_tax': city_net_tax,
                'calculated_total_net_tax': total_net_tax
            })

            # Add to tax calculations
            tax_calcs.append({
                'jurisdiction_code': item['jurisdiction_code'],
                'city_name': item['city_name'],
                'calculated_county_tax': county_tax_amount,
                'calculated_city_tax': city_tax_amount,
                'calculated_rta_tax': rta_tax_amount,
                'calculated_state_tax': state_tax_amount,
                'calculated_total_tax': calculated_total_tax,
                'calculated_city_svc_fee': city_service_fee,
                'calculated_state_svc_fee': state_service_fee,
                'calculated_state_net_tax': state_net_tax,
                'calculated_city_net_tax': city_net_tax,
                'calculated_total_net_tax': total_net_tax
            })

            # Add to totals
            total_tax_collected += total_jurisdiction_tax
            total_standard_sales += standard_sales
            total_cig_sales += cigarette_sales
            total_soda_sales += soda_sales
            total_sales += total_jurisdiction_sales

        # Add totals row to sales results
        sales_results.append({
            'jurisdiction_code': '',
            'city_name': '',
            'jurisdiction_name': '',
            'total_tax_collected': total_tax_collected,
            'standard_sales': total_standard_sales,
            'cigarette_sales': total_cig_sales,
            'soda_sales': total_soda_sales,
            'total_sales': total_sales,
            'calculated_state_net_tax': sum(item['calculated_state_net_tax'] for item in tax_calcs),
            'calculated_city_tax': sum(item['calculated_city_net_tax'] for item in tax_calcs),
            'calculated_total_net_tax': sum(item['calculated_total_net_tax'] for item in tax_calcs)
        })

        # Add totals row to tax calculations
        tax_calcs.append({
            'jurisdiction_code': '',
            'city_name': '',
            'calculated_county_tax': sum(item['calculated_county_tax'] for item in tax_calcs),
            'calculated_city_tax': sum(item['calculated_city_tax'] for item in tax_calcs),
            'calculated_rta_tax': sum(item['calculated_rta_tax'] for item in tax_calcs),
            'calculated_state_tax': sum(item['calculated_state_tax'] for item in tax_calcs),
            'calculated_total_tax': sum(item['calculated_total_tax'] for item in tax_calcs),
            'calculated_city_svc_fee': sum(item['calculated_city_svc_fee'] for item in tax_calcs),
            'calculated_state_svc_fee': sum(item['calculated_state_svc_fee'] for item in tax_calcs),
            'calculated_state_net_tax': sum(item['calculated_state_net_tax'] for item in tax_calcs),
            'calculated_city_net_tax': sum(item['calculated_city_net_tax'] for item in tax_calcs),
            'calculated_total_net_tax': sum(item['calculated_total_net_tax'] for item in tax_calcs)
        })

        # Convert lists to DataFrames
        sales_df = pd.DataFrame(sales_results)
        tax_calcs_df = pd.DataFrame(tax_calcs)

        return {
            'sales_calculations': sales_df,
            'tax_calculations': tax_calcs_df
        }

    def generate_report(self, tax_data, file_path):
        """
        Generate an Excel report with sales tax calculations

        Args:
            tax_data: List of dictionaries containing jurisdiction codes and tax amounts
            file_path: Path where the Excel file will be saved

        Returns:
            Path to the generated Excel file
        """
        # Calculate results
        results = self.calculate_sales_from_tax(tax_data)

        # Extract the DataFrames
        sales_df = results['sales_calculations']
        tax_calcs_df = results['tax_calculations']

        # Create a new workbook
        wb = openpyxl.Workbook()

        # Add the Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Sales Tax Summary"

        # Add title
        ws_summary['A1'] = "CHEERS LIQUOR MART"
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary.merge_cells('A1:G1')
        ws_summary['A1'].alignment = Alignment(horizontal='center')

        ws_summary['A2'] = "SALES TAX CALCULATION REPORT"
        ws_summary['A2'].font = Font(size=12, bold=True)
        ws_summary.merge_cells('A2:G2')
        ws_summary['A2'].alignment = Alignment(horizontal='center')

        ws_summary['A3'] = f"Generated on: {datetime.now().strftime('%m/%d/%Y %H:%M:%S')}"
        ws_summary.merge_cells('A3:G3')
        ws_summary['A3'].alignment = Alignment(horizontal='center')

        # Add Sales Calculations section header
        ws_summary['A5'] = "Sales and Tax Summary by Jurisdiction"
        ws_summary['A5'].font = Font(bold=True)
        ws_summary.merge_cells('A5:G5')

        # Write Sales Calculations headers
        headers = [
            "Jurisdiction Code", "City Name", "Jurisdiction Name", "Total Tax Collected $",
            "Standard Sales $", "Cigarette Sales $", "Soda Sales $", "Total Sales $",
            "Calculated State Net Tax $", "Calculated City Tax $", "Calculated Total Net Tax $"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Write Sales Calculations data
        for row_idx, row_data in enumerate(sales_df.itertuples(), start=7):
            is_total_row = row_idx == (len(sales_df) + 6)

            # Write each column
            ws_summary.cell(row=row_idx, column=1).value = row_data.jurisdiction_code
            ws_summary.cell(row=row_idx, column=2).value = row_data.city_name
            ws_summary.cell(row=row_idx, column=3).value = row_data.jurisdiction_name
            ws_summary.cell(row=row_idx, column=4).value = row_data.total_tax_collected
            ws_summary.cell(row=row_idx, column=5).value = row_data.standard_sales
            ws_summary.cell(row=row_idx, column=6).value = row_data.cigarette_sales
            ws_summary.cell(row=row_idx, column=7).value = row_data.soda_sales
            ws_summary.cell(row=row_idx, column=8).value = row_data.total_sales
            ws_summary.cell(row=row_idx, column=9).value = row_data.calculated_state_net_tax
            ws_summary.cell(row=row_idx, column=10).value = row_data.calculated_city_tax
            ws_summary.cell(row=row_idx, column=11).value = row_data.calculated_total_net_tax

            # Format money cells
            for col in [4, 5, 6, 7, 8, 9, 10, 11]:
                cell = ws_summary.cell(row=row_idx, column=col)
                cell.number_format = '$#,##0.00'

                # Add bold formatting for total row
                if row_idx == (len(sales_df) + 6):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Add bold formatting for total row text cells
            if row_idx == (len(sales_df) + 6):
                for col in [1, 2, 3]:
                    cell = ws_summary.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Add Tax Calculations section
        start_row = len(sales_df) + 8

        ws_summary.cell(row=start_row, column=1).value = "Detailed Tax Calculations by Jurisdiction"
        ws_summary.cell(row=start_row, column=1).font = Font(bold=True)
        ws_summary.merge_cells(f'A{start_row}:G{start_row}')

        # Write Tax Calculations headers
        headers = [
            "Jurisdiction Code", "City Name", "County Tax $", "City Tax $", "RTA Tax $",
            "State Tax $", "Total Tax $", "City Service Fee $", "State Service Fee $",
            "State Net Tax $", "City Net Tax $", "Total Net Tax $"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=start_row+1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Write Tax Calculations data
        for row_idx, row_data in enumerate(tax_calcs_df.itertuples(), start=start_row+2):
            is_total_row = row_idx == (start_row+2+len(tax_calcs_df)-1)

            # Write each column
            ws_summary.cell(row=row_idx, column=1).value = row_data.jurisdiction_code
            ws_summary.cell(row=row_idx, column=2).value = row_data.city_name
            ws_summary.cell(row=row_idx, column=3).value = row_data.calculated_county_tax
            ws_summary.cell(row=row_idx, column=4).value = row_data.calculated_city_tax
            ws_summary.cell(row=row_idx, column=5).value = row_data.calculated_rta_tax
            ws_summary.cell(row=row_idx, column=6).value = row_data.calculated_state_tax
            ws_summary.cell(row=row_idx, column=7).value = row_data.calculated_total_tax
            ws_summary.cell(row=row_idx, column=8).value = row_data.calculated_city_svc_fee
            ws_summary.cell(row=row_idx, column=9).value = row_data.calculated_state_svc_fee
            ws_summary.cell(row=row_idx, column=10).value = row_data.calculated_state_net_tax
            ws_summary.cell(row=row_idx, column=11).value = row_data.calculated_city_net_tax
            ws_summary.cell(row=row_idx, column=12).value = row_data.calculated_total_net_tax

            # Format money cells
            for col in range(3, 13):
                cell = ws_summary.cell(row=row_idx, column=col)
                cell.number_format = '$#,##0.00'

                # Add bold formatting for total row
                if is_total_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Add bold formatting for total row text cells
            if is_total_row:
                for col in [1, 2]:
                    cell = ws_summary.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Add Tax Rates Reference section
        start_row = start_row + len(tax_calcs_df) + 3

        ws_summary.cell(row=start_row, column=1).value = "Tax Rates Reference"
        ws_summary.cell(row=start_row, column=1).font = Font(bold=True)
        ws_summary.merge_cells(f'A{start_row}:G{start_row}')

        # Write Tax Rates headers
        headers = [
            "Jurisdiction Code", "City Name", "Jurisdiction Name", "County", "City Collection",
            "City Tax %", "County Tax %", "RTA Tax %", "State Tax %", "Total Tax %",
            "State Service Fee %", "City Service Fee %"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=start_row+1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Write Tax Rates data
        row_idx = start_row + 2
        for code, data in self.jurisdictions.items():
            ws_summary.cell(row=row_idx, column=1).value = code.split('-')[0]  # Remove suffixes for display
            ws_summary.cell(row=row_idx, column=2).value = data['city_name']
            ws_summary.cell(row=row_idx, column=3).value = data['name']
            ws_summary.cell(row=row_idx, column=4).value = data['county']
            ws_summary.cell(row=row_idx, column=5).value = data['city_collection']
            ws_summary.cell(row=row_idx, column=6).value = data['city_tax']
            ws_summary.cell(row=row_idx, column=7).value = data['county_tax']
            ws_summary.cell(row=row_idx, column=8).value = data['rta_tax']
            ws_summary.cell(row=row_idx, column=9).value = data['state_tax']
            ws_summary.cell(row=row_idx, column=10).value = data['total_tax']
            ws_summary.cell(row=row_idx, column=11).value = data['state_service_fee']
            ws_summary.cell(row=row_idx, column=12).value = data['city_service_fee']

            # Format percentage cells
            for col in range(6, 13):
                cell = ws_summary.cell(row=row_idx, column=col)
                cell.number_format = '0.00%'

            row_idx += 1

        # Set column widths
        column_widths = {
            'A': 20, 'B': 25, 'C': 25, 'D': 22, 'E': 22,
            'F': 22, 'G': 22, 'H': 22, 'I': 22, 'J': 22,
            'K': 22, 'L': 22
        }

        for col_letter, width in column_widths.items():
            ws_summary.column_dimensions[col_letter].width = width

        # Create an Input Data sheet
        ws_input = wb.create_sheet(title="Input Data")

        # Add title
        ws_input['A1'] = "CHEERS LIQUOR MART"
        ws_input['A1'].font = Font(size=14, bold=True)
        ws_input.merge_cells('A1:E1')
        ws_input['A1'].alignment = Alignment(horizontal='center')

        ws_input['A2'] = "SALES TAX INPUT DATA"
        ws_input['A2'].font = Font(size=12, bold=True)
        ws_input.merge_cells('A2:E2')
        ws_input['A2'].alignment = Alignment(horizontal='center')

        # Add Input Data headers
        headers = [
            "Jurisdiction Code", "City Name", "Standard Tax Collected $",
            "Cigarette Tax Collected $", "Soda Tax Collected $"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_input.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Write Input Data
        for row_idx, data in enumerate(tax_data, start=5):
            ws_input.cell(row=row_idx, column=1).value = data.get('jurisdiction_code', '')
            ws_input.cell(row=row_idx, column=2).value = data.get('city_name', '')
            ws_input.cell(row=row_idx, column=3).value = float(data.get('standard_tax', 0) or 0)
            ws_input.cell(row=row_idx, column=4).value = float(data.get('cigarette_tax', 0) or 0)
            ws_input.cell(row=row_idx, column=5).value = float(data.get('soda_tax', 0) or 0)

            # Format money cells
            for col in [3, 4, 5]:
                cell = ws_input.cell(row=row_idx, column=col)
                cell.number_format = '$#,##0.00'

        # Set column widths
        column_widths = {'A': 20, 'B': 25, 'C': 25, 'D': 25, 'E': 25}
        for col_letter, width in column_widths.items():
            ws_input.column_dimensions[col_letter].width = width

        # Save the workbook
        wb.save(file_path)

        return file_path
