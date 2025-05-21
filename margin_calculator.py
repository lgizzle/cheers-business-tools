import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.axis import ChartLines
import os
import json
from datetime import datetime

class MarginCalculator:
    """
    Calculator for determining pricing, margins, and markups for retail products.
    Includes sensitivity analysis for different margin levels.
    """
    def __init__(self):
        # Default margin range for sensitivity analysis
        self.min_margin = 0.23  # 23%
        self.max_margin = 0.35  # 35%
        self.step_margin = 0.01  # 1% steps

    def set_margin_range(self, min_margin, max_margin, step_margin=0.01):
        """Set the margin range for sensitivity analysis"""
        self.min_margin = float(min_margin)
        self.max_margin = float(max_margin)
        self.step_margin = float(step_margin)

    def calculate_price_from_margin(self, cost, margin):
        """
        Calculate selling price based on cost and desired margin
        Margin = (Price - Cost) / Price
        Price = Cost / (1 - Margin)
        """
        try:
            cost = float(cost)
            margin = float(margin)

            if margin >= 1:
                return None  # Margin must be less than 100%

            price = cost / (1 - margin)
            return price
        except (ValueError, ZeroDivisionError):
            return None

    def calculate_price_from_markup(self, cost, markup):
        """
        Calculate selling price based on cost and desired markup
        Markup = (Price - Cost) / Cost
        Price = Cost * (1 + Markup)
        """
        try:
            cost = float(cost)
            markup = float(markup)

            price = cost * (1 + markup)
            return price
        except ValueError:
            return None

    def calculate_margin_from_price(self, cost, price):
        """
        Calculate margin given cost and price
        Margin = (Price - Cost) / Price
        """
        try:
            cost = float(cost)
            price = float(price)

            if price <= 0:
                return None

            margin = (price - cost) / price
            return margin
        except (ValueError, ZeroDivisionError):
            return None

    def calculate_markup_from_price(self, cost, price):
        """
        Calculate markup given cost and price
        Markup = (Price - Cost) / Cost
        """
        try:
            cost = float(cost)
            price = float(price)

            if cost <= 0:
                return None

            markup = (price - cost) / cost
            return markup
        except (ValueError, ZeroDivisionError):
            return None

    def perform_sensitivity_analysis(self, cost, current_price=None):
        """
        Perform sensitivity analysis across a range of margins
        Returns a DataFrame with pricing at different margin levels
        """
        try:
            cost = float(cost)
            if current_price:
                current_price = float(current_price)

            # Generate margin range
            margins = np.arange(self.min_margin, self.max_margin + self.step_margin, self.step_margin)

            # Calculate for each margin
            results = []
            for margin in margins:
                price = self.calculate_price_from_margin(cost, margin)
                markup = self.calculate_markup_from_price(cost, price)
                profit = price - cost

                results.append({
                    'margin': margin,
                    'markup': markup,
                    'price': price,
                    'profit': profit,
                    'is_current': False if current_price is None else
                                 np.isclose(price, current_price, rtol=1e-3)
                })

            # Create DataFrame
            df = pd.DataFrame(results)

            # Add current price analysis if provided
            if current_price:
                current_margin = self.calculate_margin_from_price(cost, current_price)
                current_markup = self.calculate_markup_from_price(cost, current_price)
                current_profit = current_price - cost

                df = df.append({
                    'margin': current_margin,
                    'markup': current_markup,
                    'price': current_price,
                    'profit': current_profit,
                    'is_current': True
                }, ignore_index=True)

                # Sort by margin
                df = df.sort_values('margin')

            return df
        except ValueError:
            return None

    def perform_batch_analysis(self, products_data):
        """
        Analyze multiple products at once
        products_data: List of dicts with 'name', 'cost', and optionally 'current_price'
        target_margin: Target margin to apply
        """
        results = []

        for product in products_data:
            name = product.get('name', 'Unnamed')
            cost = float(product.get('cost', 0))
            current_price = product.get('current_price')

            if current_price:
                current_price = float(current_price)
                current_margin = self.calculate_margin_from_price(cost, current_price)
                current_markup = self.calculate_markup_from_price(cost, current_price)
                current_profit = current_price - cost
            else:
                current_margin = None
                current_markup = None
                current_profit = None

            # Target pricing at different margins
            pricing_options = {}
            for margin in np.arange(0.20, 0.41, 0.05):  # 20%, 25%, 30%, 35%, 40%
                price = self.calculate_price_from_margin(cost, margin)
                pricing_options[f"price_at_{int(margin*100)}pct"] = price
                pricing_options[f"profit_at_{int(margin*100)}pct"] = price - cost

            results.append({
                'name': name,
                'cost': cost,
                'current_price': current_price,
                'current_margin': current_margin,
                'current_markup': current_markup,
                'current_profit': current_profit,
                **pricing_options
            })

        # Convert to DataFrame
        return pd.DataFrame(results)

    def generate_report(self, data, file_path):
        """
        Generate an Excel report with margin calculations and sensitivity analysis

        Args:
            data: Dict containing calculation data
            file_path: Path to save the Excel file

        Returns:
            Path to the generated Excel file
        """
        # Create workbook
        wb = openpyxl.Workbook()

        # Product Details Sheet
        ws_details = wb.active
        ws_details.title = "Product Details"

        # Set up Product Details sheet
        self._generate_details_sheet(ws_details, data)

        # Sensitivity Analysis Sheet
        ws_sensitivity = wb.create_sheet("Sensitivity Analysis")

        # Set up Sensitivity Analysis sheet
        self._generate_sensitivity_sheet(ws_sensitivity, data)

        # Add charts if sensitivity data is available
        if 'sensitivity_data' in data and data['sensitivity_data'] is not None:
            # Create charts sheet
            ws_charts = wb.create_sheet("Charts")
            self._generate_charts_sheet(ws_charts, data)

        # Save the workbook
        wb.save(file_path)

        return file_path

    def _generate_details_sheet(self, ws, data):
        """Generate the Product Details sheet"""
        # Add title
        ws['A1'] = "MARGIN/MARKUP CALCULATOR"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add product details
        ws['A3'] = "Product Name:"
        ws['B3'] = data.get('product_name', 'N/A')
        ws['A3'].font = Font(bold=True)

        ws['A4'] = "Product Cost:"
        ws['B4'] = data.get('cost', 0)
        ws['B4'].number_format = '$#,##0.00'
        ws['A4'].font = Font(bold=True)

        if 'current_price' in data and data['current_price']:
            ws['A5'] = "Current Price:"
            ws['B5'] = data['current_price']
            ws['B5'].number_format = '$#,##0.00'
            ws['A5'].font = Font(bold=True)

        # Add calculation results
        ws['A7'] = "CALCULATION RESULTS"
        ws['A7'].font = Font(size=14, bold=True)
        ws.merge_cells('A7:E7')

        # Margin calculations
        ws['A9'] = "Target Margin:"
        ws['B9'] = data.get('target_margin', 0)
        ws['B9'].number_format = '0.0%'
        ws['A9'].font = Font(bold=True)

        ws['A10'] = "Price at Target Margin:"
        ws['B10'] = data.get('price_at_target_margin', 0)
        ws['B10'].number_format = '$#,##0.00'
        ws['A10'].font = Font(bold=True)

        ws['A11'] = "Profit at Target Margin:"
        ws['B11'] = data.get('price_at_target_margin', 0) - data.get('cost', 0)
        ws['B11'].number_format = '$#,##0.00'
        ws['A11'].font = Font(bold=True)

        # Markup calculations
        ws['D9'] = "Equivalent Markup:"
        markup = data.get('target_margin', 0) / (1 - data.get('target_margin', 0))
        ws['E9'] = markup
        ws['E9'].number_format = '0.0%'
        ws['D9'].font = Font(bold=True)

        # If current price provided, show current margin & markup
        if 'current_price' in data and data['current_price']:
            ws['A13'] = "CURRENT PRICING ANALYSIS"
            ws['A13'].font = Font(size=14, bold=True)
            ws.merge_cells('A13:E13')

            current_margin = data.get('current_margin', 0)
            current_markup = data.get('current_markup', 0)

            ws['A15'] = "Current Margin:"
            ws['B15'] = current_margin
            ws['B15'].number_format = '0.0%'
            ws['A15'].font = Font(bold=True)

            ws['D15'] = "Current Markup:"
            ws['E15'] = current_markup
            ws['E15'].number_format = '0.0%'
            ws['D15'].font = Font(bold=True)

            ws['A16'] = "Current Profit:"
            ws['B16'] = data.get('current_price', 0) - data.get('cost', 0)
            ws['B16'].number_format = '$#,##0.00'
            ws['A16'].font = Font(bold=True)

            # Difference from target
            profit_diff = (data.get('price_at_target_margin', 0) - data.get('cost', 0)) - (data.get('current_price', 0) - data.get('cost', 0))

            ws['A18'] = "Profit Difference (Target vs Current):"
            ws['B18'] = profit_diff
            ws['B18'].number_format = '$#,##0.00'
            ws['A18'].font = Font(bold=True)

            # Apply conditional formatting
            if profit_diff > 0:
                ws['B18'].font = Font(bold=True, color="006100")
            elif profit_diff < 0:
                ws['B18'].font = Font(bold=True, color="9C0006")

        # Set column widths
        for col, width in [('A', 30), ('B', 15), ('C', 5), ('D', 20), ('E', 15)]:
            ws.column_dimensions[col].width = width

    def _generate_sensitivity_sheet(self, ws, data):
        """Generate the Sensitivity Analysis sheet"""
        # Add title
        ws['A1'] = "MARGIN SENSITIVITY ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add product info
        ws['A3'] = "Product Name:"
        ws['B3'] = data.get('product_name', 'N/A')
        ws['A3'].font = Font(bold=True)

        ws['A4'] = "Product Cost:"
        ws['B4'] = data.get('cost', 0)
        ws['B4'].number_format = '$#,##0.00'
        ws['A4'].font = Font(bold=True)

        # Check if sensitivity data exists
        sensitivity_data = data.get('sensitivity_data')
        if sensitivity_data is not None and not sensitivity_data.empty:
            # Add table headers
            headers = ['Margin', 'Markup', 'Price', 'Profit', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True, color="FFFFFF")

            # Add data rows
            for i, row in enumerate(sensitivity_data.itertuples(), 7):
                ws.cell(row=i, column=1).value = row.margin
                ws.cell(row=i, column=1).number_format = '0.0%'

                ws.cell(row=i, column=2).value = row.markup
                ws.cell(row=i, column=2).number_format = '0.0%'

                ws.cell(row=i, column=3).value = row.price
                ws.cell(row=i, column=3).number_format = '$#,##0.00'

                ws.cell(row=i, column=4).value = row.profit
                ws.cell(row=i, column=4).number_format = '$#,##0.00'

                # Add notes for current price and sweet spot
                notes = []
                if hasattr(row, 'is_current') and row.is_current:
                    notes.append("Current Price")
                    for j in range(1, 5):
                        cell = ws.cell(row=i, column=j)
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

                # Highlight "sweet spot" range (28-32%)
                if 0.28 <= row.margin <= 0.32:
                    notes.append("Sweet Spot")
                    for j in range(1, 5):
                        cell = ws.cell(row=i, column=j)
                        if not (hasattr(row, 'is_current') and row.is_current):
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

                ws.cell(row=i, column=5).value = ", ".join(notes) if notes else ""
        else:
            ws['A6'] = "No sensitivity data available"

        # Set column widths
        for col, width in [('A', 10), ('B', 10), ('C', 12), ('D', 12), ('E', 20)]:
            ws.column_dimensions[col].width = width

    def _generate_charts_sheet(self, ws, data):
        """Generate charts for visual analysis"""
        # Add title
        ws['A1'] = "PRICING ANALYSIS CHARTS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:I1')
        ws['A1'].alignment = Alignment(horizontal='center')

        sensitivity_data = data.get('sensitivity_data')
        if sensitivity_data is None or sensitivity_data.empty:
            ws['A3'] = "No data available for charts"
            return

        # Create a table of data for the charts
        ws['A3'] = "Margin"
        ws['B3'] = "Price"
        ws['C3'] = "Profit"

        for i, row in enumerate(sensitivity_data.itertuples(), 4):
            ws.cell(row=i, column=1).value = row.margin
            ws.cell(row=i, column=1).number_format = '0.0%'

            ws.cell(row=i, column=2).value = row.price
            ws.cell(row=i, column=2).number_format = '$#,##0.00'

            ws.cell(row=i, column=3).value = row.profit
            ws.cell(row=i, column=3).number_format = '$#,##0.00'

        # Create Margin vs Price Line Chart
        chart1 = LineChart()
        chart1.title = "Margin vs Price"
        chart1.style = 2
        chart1.x_axis.title = "Margin"
        chart1.y_axis.title = "Price"

        # Define data ranges
        data_rows = len(sensitivity_data) + 1  # +1 for header
        x_values = Reference(ws, min_col=1, min_row=3, max_row=3+data_rows)
        y_values = Reference(ws, min_col=2, min_row=3, max_row=3+data_rows)

        chart1.add_data(y_values, titles_from_data=True)
        chart1.set_categories(x_values)

        # Place chart
        ws.add_chart(chart1, "E3")

        # Create Margin vs Profit Bar Chart
        chart2 = BarChart()
        chart2.title = "Margin vs Profit"
        chart2.style = 2
        chart2.x_axis.title = "Margin"
        chart2.y_axis.title = "Profit"

        y_values = Reference(ws, min_col=3, min_row=3, max_row=3+data_rows)

        chart2.add_data(y_values, titles_from_data=True)
        chart2.set_categories(x_values)

        # Place chart
        ws.add_chart(chart2, "E20")

        # Set column widths
        for col, width in [('A', 10), ('B', 12), ('C', 12)]:
            ws.column_dimensions[col].width = width
