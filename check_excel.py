import pandas as pd
import openpyxl

try:
    wb = openpyxl.load_workbook('Deal Split Calculator V01-3.xlsm', read_only=True, keep_vba=True)
    print('Sheets in workbook:', wb.sheetnames)

    # Examine Instructions sheet
    print("\nExamining Instructions sheet:")
    sheet = wb['Instructions']
    print(f"Dimensions: {sheet.min_row}-{sheet.max_row} rows, {sheet.min_column}-{sheet.max_column} columns")

    for row in range(1, min(20, sheet.max_row + 1)):
        row_values = [str(sheet.cell(row=row, column=col).value) for col in range(1, min(5, sheet.max_column + 1))]
        print(f"Row {row}: {' | '.join(row_values)}")

    # Examine DealCalc sheet
    sheet = wb['DealCalc']
    print("\nExamining DealCalc sheet:")
    print(f"Dimensions: {sheet.min_row}-{sheet.max_row} rows, {sheet.min_column}-{sheet.max_column} columns")

    # Try to read some key cells to understand structure
    print("\nTop part of DealCalc sheet:")
    for row in range(1, 16):
        row_values = [str(sheet.cell(row=row, column=col).value) for col in range(1, min(10, sheet.max_column + 1))]
        print(f"Row {row}: {' | '.join(row_values)}")

    print("\nBottom part of DealCalc sheet (calculations):")
    for row in range(20, 29):
        row_values = [str(sheet.cell(row=row, column=col).value) for col in range(1, min(10, sheet.max_column + 1))]
        print(f"Row {row}: {' | '.join(row_values)}")

    # Examine ScenarioStorage sheet - first 5 rows to understand structure
    sheet = wb['ScenarioStorage']
    print("\nExamining ScenarioStorage sheet (first 5 rows):")
    print(f"Dimensions: {sheet.min_row}-{sheet.max_row} rows, {sheet.min_column}-{sheet.max_column} columns")
    for row in range(1, min(6, sheet.max_row + 1)):
        row_values = [str(sheet.cell(row=row, column=col).value) for col in range(1, 4)]
        print(f"Row {row}: {' | '.join(row_values)}")

    # Examine Scenarios sheet - first 5 columns of first 5 rows to understand structure
    sheet = wb['Scenarios']
    print("\nExamining Scenarios sheet (first rows/columns):")
    print(f"Dimensions: {sheet.min_row}-{sheet.max_row} rows, {sheet.min_column}-{sheet.max_column} columns")
    for row in range(1, min(4, sheet.max_row + 1)):
        row_values = [str(sheet.cell(row=row, column=col).value) for col in range(1, min(15, sheet.max_column + 1))]
        print(f"Row {row}: {' | '.join(row_values)}")

    # Summarize what we've learned about the Deal Split Calculator
    print("\n--- DEAL SPLIT CALCULATOR SUMMARY ---")
    print("1. The tool calculates how to split orders based on annual sales proportions.")
    print("2. DealCalc sheet: Main calculation sheet where users enter variety names, annual sales, and desired total order.")
    print("3. ScenarioStorage sheet: Stores individual variety sales figures.")
    print("4. Scenarios sheet: Stores complete scenarios with variety names and their annual sales.")
    print("5. Key formulas: C4*$C$26/$C$24 (calculates split based on proportion of total sales)")

except Exception as e:
    print(f'Error: {e}')
